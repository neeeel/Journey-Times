import tkinter
import tkinter.ttk as ttk
import tkinter.font as font
import datetime
import mapmanager
import openpyxl
import win32com.client
from PIL import Image,ImageDraw,ImageTk,ImageFont
from tkinter import filedialog
from tkinter import messagebox
import threading
import queue
import copy
import time
import os
import dragandzoomcanvas
import mapmanager2
import pandas as pd
import mapViewer
import pickle
import shutil





class mainWindow(tkinter.Tk):

    def __init__(self):
        super(mainWindow, self).__init__()
        self.routes = {}
        self.tracsisBlue = "#%02x%02x%02x" % (20, 27, 77)
        ttk.Style().configure(".",foreground = self.tracsisBlue,background = "white",weight="bold")
        #self.bind("<<finished>>",self.received)
        self.selectedRoute = None
        self.progress = None
        self.trackData = None
        self.mapViewer = None
        self.unitsVar = tkinter.IntVar()
        self.q = queue.Queue()
        self.discardedRuns = []
        self.baseData = [] ### store the returned data, ready to display either normal or discarded runs
        self.normalRuns=[]
        self.colours = ["black","red","gold","sea green","deep sky blue"]
        self.processSingleTrackFunction = None
        self.loadRoutesFunction = None
        self.getDateFunction = None
        self.getSpeedFunction = None
        self.getfullDataframeFunction = None
        self.previousLegIndex = -1 # this keeps track of the leg that is displayed yellow, so we dont have to redraw the whole track each time
        self.primaryTrackList = []
        self.secondaryTrackList = []
        self.trackWindow = None
        self.mapMan = None
        self.mapImage = None
        self.addTimingPointFlag = False
        self.wm_title("JoPro - Journey Time Software")
        self.state("zoomed")
        self.primaryTrees = []
        self.secondaryTrees = []
        self.trees = [None,None]
        self.configure(bg="white")

        ###
        ### set up menu bar
        ###
        self.menubar = tkinter.Menu(self)
        menu = tkinter.Menu(self.menubar,tearoff = 0)
        menu.add_command(label = "Load TPs",command = self.displayRoutes)
        menu.add_separator()
        menu.add_command(label="Export",command=self.spawn_export_window)
        menu.add_separator()
        menu.add_command(label="View Full Track", command=self.view_full_track)
        self.menubar.add_cascade(label="File",menu = menu)
        self.config(menu=self.menubar)
        menu = tkinter.Menu(self.menubar,tearoff = 0)
        menu.add_command(label ="Journey Time Settings",command = self.spawn_settings_window)
        #menu.add_separator()
        #menu.add_command(label = "Excel Settings",command = self.spawn_excel_window)
        self.menubar.add_cascade(label = "Settings",menu = menu)
        self.thumbnail = None


        leftFrame = tkinter.Frame(self, width=100, height=900, bg="white",relief=tkinter.GROOVE,borderwidth=2)
        labelFont = font.Font(family='Helvetica', size=16, weight='bold')
        tkinter.Label(leftFrame,text = "Route",font=labelFont,bg=self.tracsisBlue,fg="white",relief=tkinter.GROOVE,borderwidth=2).grid(row=0,column=0,sticky="nsew")
        self.routeListBox = tkinter.Listbox(leftFrame,height = 10,width=17,relief=tkinter.GROOVE,borderwidth=2,bg="white")
        self.routeListBox.grid(row=1,column=0,pady=(20,0))
        self.routeListBox.bind('<Double-Button-1>', lambda event: self.startProgress("Calculating Runs",self.runProcess))
        self.routeListBox.bind('<<ListboxSelect>>', self.showTPs)


        leftFrame.grid(row=0,column=0,sticky="n")
        mapFrame = tkinter.Frame(self,width=800,height=800,relief=tkinter.GROOVE,borderwidth=2)
        mapFrame.grid(row=0,column=1)
        tkinter.Label(mapFrame,text="Map",font=labelFont,bg=self.tracsisBlue,fg="white",relief=tkinter.GROOVE,borderwidth=2).grid(row=0,column=0,sticky="nsew")
        self.thumbnailCanvas = tkinter.Canvas(mapFrame,width = 800,height = 800,relief=tkinter.GROOVE,borderwidth =2,bg="white")
        self.thumbnailCanvas.grid(row=1, column=0,pady=(20,0), sticky="nw")
        tkinter.Label(mapFrame,text="Journey Time Summary",font=labelFont,bg=self.tracsisBlue,fg="white",relief=tkinter.GROOVE,borderwidth=2).grid(row=0,column=1,sticky="nsew")
        self.tabs = ttk.Notebook(mapFrame)
        self.primaryFrame = tkinter.Frame(mapFrame, relief=tkinter.GROOVE,borderwidth=2, width=800, height=800, bg="white")
        self.secondaryFrame = tkinter.Frame(mapFrame, relief=tkinter.GROOVE,borderwidth=2, width=800, height=800, bg="white")
        self.tabs.add(self.primaryFrame, text="Primary")
        self.tabs.add(self.secondaryFrame, text="Secondary")
        self.tabs.bind("<<NotebookTabChanged>>", self.tabChanged)
        self.tabs.grid(row=1, column=1, sticky="nw")
        print(self.tabs.tabs())
        frame = self.nametowidget(self.tabs.tabs()[0])
        print("frame is",type(frame))
        return

    ######################################################################################################
    #
    # Stuff to deal with loading and displaying routes
    #
    ######################################################################################################

    def showTPs(self,event):
        ### Get the current selection in the routes List box, and display the TPS for that route in the TPS list box
        ###
        self.thumbnailCanvas.delete(tkinter.ALL)
        label = self.winfo_children()[2].winfo_children()[0]
        routeName = self.routeListBox.get(self.routeListBox.curselection())
        if routeName in self.routes:
            self.thumbnailCanvas.delete(tkinter.ALL)
            label.configure(text="Map - " + routeName)
            if self.routes[routeName].changed:
                self.load_route_maps(self.routes[routeName])
                self.routes[routeName].changed = False
            self.thumbnail = ImageTk.PhotoImage(self.routes[routeName].get_primary_map())
            self.thumbnailCanvas.create_image(0, 0, image=self.thumbnail, anchor=tkinter.NW)
            self.mapMan = self.routes[routeName].getMapManager()

    def displayRoutes(self):
        for frame in self.tabs.tabs():
            frame = self.nametowidget(frame)
            for child in frame.winfo_children():
                child.destroy()
        self.routeListBox.delete(0,tkinter.END)
        self.routes = self.loadRoutesFunction()
        if self.routes is None:
            return
        for k,v in self.routes.items():
            self.load_route_maps(v)
            self.routeListBox.insert(tkinter.END,v.name)

    def load_route_maps(self,route):
        print("setting up maps for ",route.name)
        prim = [(x[2], x[3]) for x in route.get_timing_points()[0]]
        sec = [(x[2], x[3]) for x in route.get_timing_points()[1]]
        mp = mapmanager.MapManager(640, 640, 11, [], [prim, sec])
        prim, sec = mp.get_thumbnails()
        #prim.show()
        route.add_primary_map(prim)
        route.add_secondary_map(sec)
        route.setMapManager(mp)

    ######################################################################################################
    #
    # Stuff to deal with progress bar
    #
    ######################################################################################################

    def startProgress(self,msg,fun):
        threading.Thread(target=fun).start()
        self.progressWin = tkinter.Toplevel(self,width = 200,height = 200)
        x = int(self.winfo_screenwidth()/2 - 100)
        y = int(self.winfo_screenheight() / 2 - 100)
        self.progressWin.attributes("-topmost",True)
        tkinter.Label(self.progressWin,text = msg).grid(row=0,column = 0,padx = 20,pady= 20)
        self.progress = ttk.Progressbar(self.progressWin, orient="horizontal", length=200, mode="indeterminate")
        self.progress.grid(row=1,column = 0,padx = 20,pady= 20)
        self.progress.start(10)
        self.progressWin.geometry("+" + str(x) + "+" + str(y))
        self.progressWin.lift()
        self.processing = True

    def step_progress(self):
        while self.processing == True:
            self.progress.step()
            self.after(100,self.step_progress)

    def stopProgress(self):
        #print("in self.stop progress")
        self.processing = False
        if self.progress is None:
            pass
        else:
            self.progress.stop()
            self.progress =None
        self.progressWin.destroy()

    ######################################################################################################
    #
    #
    #
    ######################################################################################################

    def spawn_export_window(self):
        labelFont = font.Font(family='Times', size=12)
        self.excelWindow = tkinter.Toplevel(self)
        self.excelWindow.protocol("WM_DELETE_WINDOW", self.excel_settings_closed)
        frame = tkinter.Frame(self.excelWindow)
        frame.grid(row=0,column=0)
        ###
        ### read times from settings
        ###
        times = []
        with open('settings.txt', 'r') as f:
            for i in range(6):
                text = f.readline()
                text = text.replace("\n", "")
                times.append(text)
        # routeName = self.routeListBox.get(tkinter.ACTIVE)
        print(";",self.routeListBox.curselection(),":")
        if len(self.routeListBox.curselection()) > 0:
            self.routeListBox.activate(self.routeListBox.curselection())
        self.routeListBox.config(state=tkinter.DISABLED)
        for row,text in enumerate(["AM Times","IP Times","PM Times"]):
            tkinter.Label(frame,text=text,font=labelFont, relief=tkinter.FLAT,borderwidth=1,anchor="e",width = 14).grid(row=row,column=0,sticky="nsew")
            e = tkinter.Entry(frame,width=7)
            e.grid(row=row,column=1)
            e.insert(0,times[(row*2)])
            e.bind("<FocusOut>", self.validate_time)
            tkinter.Label(frame, text="To",font=labelFont, relief=tkinter.FLAT,borderwidth=1,width = 4).grid(row=row, column=2, sticky="nsew")
            e = tkinter.Entry(frame, width=7)
            e.grid(row=row, column=3,padx=(0,20))
            e.insert(0, times[(row * 2)+1])
            e.bind("<FocusOut>", self.validate_time)
        tkinter.Label(frame,text="Primary Dir",font=labelFont, relief=tkinter.FLAT,borderwidth=1,anchor="e",width = 14).grid(row=3,column=0,sticky="nsew")
        ttk.Combobox(frame, values=["North", "South", "East", "West", "Clockwise", "Anticlockwise"],width=13).grid(row=3, column=1,columnspan=3)
        var = tkinter.IntVar()
        var.set(1)
        c = tkinter.Checkbutton(frame, text="Primary",variable =var)
        c.grid(row=4, column=0)
        c.var = var
        var = tkinter.IntVar()
        var.set(1)
        c = tkinter.Checkbutton(frame, text="Secondary",variable =var)
        c.grid(row=4, column=1,columnspan=3)
        c.var = var
        var = tkinter.IntVar()
        c = tkinter.Checkbutton(frame, text="Export Raw Data",variable =var)
        c.grid(row=5, column=1,columnspan=3)
        c.var = var
        var = tkinter.IntVar()
        c = tkinter.Checkbutton(frame, text="Add 1 Hour to Times",variable =var)
        c.grid(row=5, column=0)
        c.var = var
        tkinter.Button(frame, text="Exit", command=self.excel_settings_closed,font=labelFont,width = 8).grid(row=6, column=0)
        tkinter.Button(frame, text="Export", command=self.export,font=labelFont,width = 8).grid(row=6, column=1,columnspan=3)


    def validate_time(self,event):
        entry = event.widget
        text = entry.get()
        if len(text) == 5:
            if text[2] == ":":
                h,m = text.split(":")
                try:
                    h = int(h)
                    m =int(m)
                    if h >= 0 and h <=23 and m >=0 and m <=59:
                        return
                except Exception as e:
                    pass
        messagebox.showinfo(message="invalid time")
        entry.delete(0,tkinter.END)
        entry.focus_set()

    def export(self):
        file = filedialog.asksaveasfilename()
        if file == "":
            messagebox.showinfo(message="no file name entered, exiting Export")
            self.stopProgress()
            return
        children = self.excelWindow.winfo_children()[0].winfo_children()
        primaryDirection = children[13].get()
        if primaryDirection == "North":
            primaryDirection = "Northbound"
            secondaryDirection = "Southbound"
        if primaryDirection == "South":
            primaryDirection = "Southbound"
            secondaryDirection = "Northbound"
        if primaryDirection == "East":
            primaryDirection = "Eastbound"
            secondaryDirection = "Westbound"
        if primaryDirection == "West":
            primaryDirection = "Westbound"
            secondaryDirection = "Eastbound"
        if primaryDirection == "Clockwise":
            secondaryDirection = "Anticlockwise"
        if primaryDirection == "Anticlockwise":
            secondaryDirection = "Clockwise"
        for index,val in enumerate([children[14].var.get(),children[15].var.get()]):
            if val == 1:
                self.export_to_excel(index,file + " "  + [primaryDirection,secondaryDirection][index] )
        self.excel_settings_closed()
        self.routeListBox.config(state=tkinter.NORMAL)
        self.stopProgress()

    def export_to_excel(self,index,file):
        tree = self.trees[index]
        if tree is None:
            return
        if len(tree.get_children()) == 0:
            return
        self.clear_runs_folder()
        routeName = self.routeListBox.get(tkinter.ACTIVE)
        route = self.routes[routeName]
        prim = [(x[2], x[3]) for x in route.get_timing_points()[0]]
        sec = [(x[2], x[3]) for x in route.get_timing_points()[1]]
        #self.mapMan = mapmanager.MapManager(640, 640, 12, [[], []], [prim,sec])  # cos I altered mapmanager to take both sets of timing points, its pretty messed up
        self.mapMan.get_centre_of_points([prim,sec][index]) ## a hack to reset the centre point of the map depending on whether its prim or sec tps
        children = self.excelWindow.winfo_children()[0].winfo_children()
        if children[17].var.get() == 1:
            td = datetime.timedelta(hours=1)
        else:
            td = datetime.timedelta(hours=0)
        exportRawData = False
        if children[16].var.get() == 1:
            exportRawData = True
            print("EXPORTING RAWE DATA!!!")
        runs = [[],[],[]]
        wb = openpyxl.load_workbook("Template.xlsm", keep_vba=True)
        sheets = wb.get_sheet_names()
        ###
        ### put big logo on front sheet
        ##
        sheet = wb.get_sheet_by_name(sheets[0])
        img = Image.open("tracsis Logo.jpg")
        imgSmall = img.resize((368, 130), Image.ANTIALIAS)
        excelImageSmall = openpyxl.drawing.image.Image(imgSmall)
        sheet.add_image(excelImageSmall, "B3")
        ###
        ### put little logos on the other sheets
        ###
        for sht in sheets[1:-1]:
            sheet = wb.get_sheet_by_name(sht)
            img = Image.open("tracsis Logo.jpg")
            imgSmall = img.resize((184, 65), Image.ANTIALIAS)
            excelImageSmall = openpyxl.drawing.image.Image(imgSmall)
            sheet.add_image(excelImageSmall, "A1")
        try:
            sheet = wb.get_sheet_by_name('Temp')
        except Exception as e:
            print(e)
            return
        tpCount = len(self.baseData[index][0][0][0])
        timeSettings = []
        for i in [1,3,5,7,9,11]:
            timeSettings.append(self.excelWindow.winfo_children()[0].winfo_children()[i].get())
        timeSettings = [datetime.datetime.strptime(t,"%H:%M") for t in timeSettings]
        print("timesettings are", timeSettings)
        row = 0
        exampleRun = -1
        for child in tree.get_children():
            print("child is",child)
            runIndex = int(child.replace("run_","")) -1
            exampleRun = 0# runIndex ## we want a run that we can use to make the path map
            runData = self.baseData[index][0][runIndex]
            print("run data is",runData)
            startTime = datetime.datetime.strptime(runData[1][0],"%d/%m/%Y %H:%M:%S").time()
            for t in range(0,len(timeSettings),2):

                if startTime >= timeSettings[t].time() and startTime <= timeSettings[t+1].time():
                    print("adding to row",row)
                    if exportRawData:
                        rawData = []
                        df = self.getTrack(runData[0])
                        df["accel"] = (df["legSpeed"] - df["legSpeed"].shift(-1))/df["legTime"]
                        exportSheet = wb.create_sheet("Raw Data - Track " + str(runIndex + 1))
                        df[["Record", "Time", "Lat", "Lon", "legTime", "legSpeed","accel"]].apply(
                            lambda x: rawData.append(x.tolist()), axis=1)
                        for i, d in enumerate(rawData):
                            for j, item in enumerate(d):
                                exportSheet.cell(row=i + 1, column=j + 1).value = item
                    self.save_track_as_image(runData[0],index,runIndex +1,td)
                    runs[t//2].append(runIndex)
                    for j in range(tpCount):
                            ### add on an hour
                            sheet.cell(row=2 + row, column=1 + j).value = datetime.datetime.strftime(
                                        datetime.datetime.strptime(runData[1][j] ,"%d/%m/%Y %H:%M:%S")+ td,"%H:%M:%S")
                            try:
                                t = (datetime.datetime.strptime(runData[1][j+1]  ,"%d/%m/%Y %H:%M:%S")-
                                    datetime.datetime.strptime(runData[1][j] ,"%d/%m/%Y %H:%M:%S"))
                                sheet.cell(row=2 + row, column=2 + j + tpCount).value = str(t)
                                print("value output is",sheet.cell(row=2 + row, column=2 + j + tpCount).value)
                            except Exception as e:
                                ### we will get an error when we try to calculate the duration for the last entry in the data list
                                print("error is",e)
                                t =(datetime.datetime.strptime(runData[1][-1], "%d/%m/%Y %H:%M:%S") -
                                    datetime.datetime.strptime(runData[1][0], "%d/%m/%Y %H:%M:%S"))
                                sheet.cell(row=2 + row, column=2 + j + tpCount).value = str(t)
                            sheet.cell(row=2 + row, column=2 + j + (2 * tpCount)).value = runData[2][j]
                            sheet.cell(row=2+row,column=1+tpCount).value = runData[1][0].split(" ")[0]
                    row+=1
        sheet["A1"] = len(runs[0])
        sheet["B1"] = datetime.datetime.strftime(timeSettings[0] + td,"%H:%M")
        sheet["C1"] = datetime.datetime.strftime(timeSettings[1] + td,"%H:%M")
        sheet["D1"] = len(runs[1])
        sheet["E1"] = datetime.datetime.strftime(timeSettings[2] + td,"%H:%M")
        sheet["F1"] = datetime.datetime.strftime(timeSettings[3] + td,"%H:%M")
        sheet["G1"] = len(runs[2])
        sheet["H1"] = datetime.datetime.strftime(timeSettings[4] + td,"%H:%M")
        sheet["I1"] = datetime.datetime.strftime(timeSettings[5] + td,"%H:%M")
        sheet["J1"] = tpCount
        folder = os.path.dirname(os.path.abspath(__file__))
        folder = os.path.join(folder, "Runs\\")
        sheet["L1"] = folder
        print("Runs are ",runs)
        for i,dist in enumerate(self.baseData[index][1]):
            sheet.cell(row=1, column=13 + i).value = dist
        wb.save(filename=file + ".xlsm")
        if exampleRun == -1:
            return
        trackData = self.getTrack(self.baseData[index][0][exampleRun][0])
        lats = trackData["Lat"].tolist()
        lons = trackData["Lon"].tolist()
        pathData = list(zip(lats, lons))
        print("Track data is ", pathData)
        image = self.mapMan.get_map_with_path([prim,sec][index], pathData)
        excelImage = openpyxl.drawing.image.Image(image)
        sheet = wb.get_sheet_by_name('Location - Distance')
        sheet.add_image(excelImage, "B13")
        wb.active = 0
        try:
            print("saving openpyxl workbook", file + ".xlsm")
            wb.save(file + ".xlsm")
            print("opening win32com file")
            xl = win32com.client.Dispatch("Excel.Application")
            xl.Application.Visible = True
            xlsFile = os.path.realpath(file + ".xlsx")
            print("xslsfile is", xlsFile)
            filename = file + ".xlsm"
            print("trying to open workbook", os.path.realpath(filename), xlsFile)
            time.sleep(0.5)
            wb = xl.Workbooks.Open(Filename=os.path.realpath(filename), ReadOnly=1)
            # xl.Workbooks.Open(Filename=os.path.realpath("C:/Users/NWatson/PycharmProjects/JourneyTimes/blah" + ".xlsm"), ReadOnly=1)
            xl.Application.Run("formatfile")
            wb.Save()
            wb.Close(SaveChanges=True)
            xl.Quit()
        except PermissionError as e:
            messagebox.showinfo(message="cannot save file- " + filename + " workbook is already open, please close and run export again")
        except Exception as e:
            print("couldnt save",e)

    def load_settings(self):
        self.entryValues = []
        for i in range(18):
            self.entryValues.append(tkinter.StringVar())
        with open('settings.txt', 'r') as f:
            count = 0
            for e in self.entryValues:
                text = f.readline()
                text = text.replace("\n", "")
                if text == "":
                    pass
                e.set(text)
                print("loading",text)
            self.unitsVar.set(int(f.readline().replace("\n","")))

    def excel_settings_closed(self):
        self.excelWindow.destroy()
        self.routeListBox.config(state=tkinter.NORMAL)

    def view_full_track_as_map(self,event):
        curItem = event.widget.identify_row(event.y)
        directionIndex = self.tabs.index(self.tabs.select())
        routeName = self.routeListBox.get(tkinter.ACTIVE)
        route = self.routes[routeName]
        prim = [(x[2], x[3]) for x in route.get_timing_points()[0]]
        sec = [(x[2], x[3]) for x in route.get_timing_points()[1]]
        #self.mapMan = mapmanager.MapManager(640, 640, 12, [[], []], [prim,sec])  # cos I altered mapmanager to take both sets of timing points, its pretty messed up
        self.mapMan.get_centre_of_points([prim,sec][directionIndex]) ## a hack to reset the centre point of the map depending on whether its prim or sec tps
        trackIndex = int(curItem.replace("run_", ""))
        print("selected item is",curItem,"direction is",directionIndex,"track index iS",trackIndex)
        label = self.winfo_children()[2].winfo_children()[0]
        data = self.baseData[directionIndex][0][trackIndex-1]
        self.trackImg = ImageTk.PhotoImage(self.save_track_as_image(data[0],directionIndex,trackIndex,datetime.timedelta(hours=0)))
        self.thumbnailCanvas.delete(tkinter.ALL)
        self.thumbnailCanvas.create_image(0, 0, image=self.trackImg, anchor=tkinter.NW)
        label.configure(text="Track  " + str(trackIndex))
        if not self.mapViewer is None:
            self.mapViewer.set_run_indices((data[0][0],data[0][-1]))

    def spawn_settings_window(self):
        pass

    def tabChanged(self,event):
        pass

    def delete_run_from_tree(self,event):
        item = event.widget.selection()[0]
        parent = event.widget.parent(item)
        if parent != "":
            event.widget.delete(parent)
        else:
            event.widget.delete(item)
        return

    def runProcess(self):
        for frame in self.tabs.tabs():
            frame = self.nametowidget(frame)
            for child in frame.winfo_children():
                child.destroy()
        label = self.winfo_children()[2].winfo_children()[0]
        routeName = self.routeListBox.get(tkinter.ACTIVE)
        self.routeListBox.activate(self.routeListBox.curselection())
        if self.routes[routeName].changed:
            ###
            ### reload the route maps if the TPS have been edited
            ###
            self.routes[routeName].changed = False
            self.load_route_maps(self.routes[routeName])
        fileList = list(filedialog.askopenfilenames(initialdir=dir))
        if not routeName in self.routes or fileList == []:
            label.configure(text="Map")
            #self.journeyLabel.configure(text="Journey Time Summary")
            return
        self.selectedRoute = self.routes[routeName]
        threading.Thread(target=self.fun,args = (self.routes[routeName],fileList)).start()
        #self.startProgress("Loading and Processing Data")
        label.configure(text="Map - " + routeName)
        #self.journeyLabel.configure(text="Journey Time Summary - " + routeName)
        return

    def receive_processed_data(self, result):
        ###
        ### result is a list of lists
        ### result[0] holds the primary direction runs
        ### result[1] holds the secondary direction runs
        ### each specific direction holds [data,distances,discarded runs]
        ###
        self.baseData = result
        width = 65
        for index,tab in enumerate(self.tabs.tabs()):
            frame = self.nametowidget(tab)
            print(result[index])
            #print("wibble",result[index][0][0][0])
            if len(result[index][0]) > 0:
                cols = tuple(range(len(result[index][0][0][0]) + 2))
                canvas = tkinter.Canvas(frame,width=800,height=800,scrollregion=(0,0,500,500))
                canvas.grid(row=0,column=0)

                tree = ttk.Treeview(frame, columns=cols, show="headings", height=38) ### journeytimes tree
                tree.column("#0", width=30)
                tree.tag_configure('duration',foreground = "gray")
                tree.tag_configure('speed',foreground = "gray")
                tree.tag_configure('run', background='light goldenrod yellow', foreground=self.tracsisBlue)
                tree.column(0, width=width, anchor='center')
                tree.heading(0, text='Track')
                #tree.bind("<BackSpace>", self.key_pressed)
                tree.bind("<Delete>", self.delete_run_from_tree)
                #tree.bind("<Double-Button-1>", self.spawn_track_window)
                tree.bind("<Button-3>", self.view_full_track_as_map)
                for i in range(1, len(result[index][0][0][0])+1):
                    tree.column(i, width=width, anchor='center')
                    tree.heading(i, text='TP' + str(i))
                tree.heading(cols[-1], text="Total")
                tree.column(cols[-1], width=width, anchor='center')
                self.trees[index] = tree
                canvas.create_window(0,0,window=tree,anchor=tkinter.NW,tags=["tree",])#tree.grid(row=0,column=0)
                frame.configure(width=800, height=800)
                scroll = ttk.Scrollbar(frame,orient="horizontal")
                scroll.grid(row=1,column=0,sticky="ew")
                scroll.configure(command=canvas.xview)
                canvas.configure(xscrollcommand=scroll.set)
                for i,row in enumerate(result[index][0]):
                    print("processing row",row[1])
                    durations =[(datetime.datetime.strptime(row[1][i + 1], "%d/%m/%Y %H:%M:%S") -
                                 datetime.datetime.strptime(row[1][i],"%d/%m/%Y %H:%M:%S")) for i in range(len(row[1])-1)]
                    times = [r.split(" ")[1] for r in row[1]]
                    times.append(sum(durations, datetime.timedelta()))
                    durations = ["duration",""] + durations
                    tree.insert("","end",iid="run_" + str(i+1),values=["Track " + str(i+1)] + times,tags=("run",))
                    tree.insert("run_" + str(i+1), "end", iid="duration_" + str(i+1), values=durations,tags=("duration",))
                    tree.insert("run_" + str(i+1), "end", iid="speed_" + str(i+1), values=["speed",""] + row[2],tags=("speed",))
                canvas.configure(scrollregion=(0,0,tree.winfo_reqwidth(),tree.winfo_reqheight()))
                self.update()
        self.stopProgress()

    def setCallbackFunction(self,text,fun):
        print("setting callback for ",text,fun)
        if text == "Routes":
            self.loadRoutesFunction = fun
        if text =="process":
            self.fun = fun
        if text == "getTrack":
            self.getTrack = fun
        if text == "getDate":
            self.getDateFunction = fun
        if text == "singleTrack":
            self.processSingleTrackFunction = fun
        if text == "getSpeed":
            self.getSpeedFunction = fun
        if text == "getFullTracks":
            self.getFullTracksFunction = fun
        if text == "fullDataframe":
            print("setting fun to ",fun)
            self.getFullDataframeFunction = fun

    def save_track_as_image(self, track, index, trackNo,tdOffset):
        routeName = self.routeListBox.get(tkinter.ACTIVE)
        route = self.routes[routeName]
        if index == 0:
            image = route.get_primary_map().copy()
        else:
            image = route.get_secondary_map().copy()

        fnt = ImageFont.truetype("arial", size=18)
        drawimage = ImageDraw.Draw(image)
        trackData = self.getTrack(track)
        trackData[:-1].apply(
            lambda row: self.draw_leg_on_image((row["Lat"], row["Lon"]), (row["latNext"], row["lonNext"]),
                                               row["legSpeed"], drawimage), axis=1)

        t = datetime.datetime.strftime(trackData.iloc[0]["Time"] + tdOffset, "%H:%M:%S")
        drawimage.rectangle([0, 0, 100, 50], fill="white")
        drawimage.text((10, 10), text=t, font=fnt, fill="black")
        folder = os.path.dirname(os.path.abspath(__file__))
        folder = os.path.join(folder, "Runs\\")
        image.save(folder + "/track " + str(trackNo) + ".jpg")
        return image

    def draw_leg_on_image(self,p1,p2,speed,drawImage):
        colours = [(0,0,0), (255,0,0), (255,215,0),(46,139,87), (0,191,255)]
        speeds = [0,5,15,30,50]
        if speed < 0:
            colour = "white"
        else:
            try:
               colIndex = [n for n, i in enumerate(speeds) if i > speed][0] - 1
               colour = colours[colIndex]
            except Exception as e:
                colour = "white"
        x, y = self.mapMan.get_coords(p1)
        drawImage.ellipse([x - 5, y - 5, x + 5, y + 5], fill=colour)
        x1, y1 = self.mapMan.get_coords(p2)
        drawImage.ellipse([x1 - 5, y1 - 5, x1 + 5, y1 + 5], fill=colour)
        drawImage.line([x, y, x1, y1], fill=colour, width=6)

    def clear_runs_folder(self):
        folder = os.path.dirname(os.path.abspath(__file__))
        folder = os.path.join(folder, "Runs\\")
        print("folder is",folder)
        #shutil.rmtree(folder)
        if not os.path.exists(folder):
            os.makedirs(folder)
        for fileName in os.listdir(folder):
            file_path = os.path.join(folder, fileName)
            try:
                if os.path.isfile(file_path):
                    os.remove(file_path)
            except Exception as e:
                print(e)

    ######################################################################################################
    #
    # Functions to deal with the full track window, adding and deleting TPs, etc
    #
    ######################################################################################################

    def view_full_track(self):

        df = self.getFullTracksFunction()
        if df is None:
            return
        coords = df[["Lat","Lon"]].values.tolist()
        self.centrePoint = coords[0]
        self.routeListBox.config(state=tkinter.DISABLED)
        coords = [mapmanager2.get_coords(self.centrePoint, p, 10,size=800) for p in coords]
        win = tkinter.Toplevel()
        win.protocol("WM_DELETE_WINDOW", lambda w = win:self.full_track_view_closed(w))

        frame = tkinter.Frame(win, bg="white")
        self.trackTabs = ttk.Notebook(frame)
        self.timingPointsFrame = tkinter.Frame(frame,bg="white")
        self.trackTabs.add(self.timingPointsFrame,text = "Timing Points")
        self.trackFrame = tkinter.Frame(frame, bg="white",width = 300)
        self.trackTabs.add(self.trackFrame, text="Full Track")
        self.failedRunsFrame = tkinter.Frame(frame, bg="white")
        self.trackTabs.add(self.failedRunsFrame, text="Failed Runs")
        self.trackTabs.grid(row=0,column=0,sticky = "nsew")
        frame.pack(expand=tkinter.YES,fill=tkinter.BOTH,side=tkinter.LEFT)

        ###
        ### set up the full track treeview
        ###

        tree = ttk.Treeview(self.trackFrame,columns=[0,1,2,3,4], show="headings", height=45)
        #scroll = tkinter.Scrollbar(self.trackFrame, orient="vertical", command=tree.yview)
        self.scale = ttk.Scale(self.trackFrame,orient="vertical",command = self.scroll_with_scale,from_=0,to=len(df)-1)
        self.scale.grid(row=0,column=1,sticky="ns")
        for index,heading in enumerate([("Index",40),("Track",60),("Lat",80),("Lon",80),("Time",120)]):
            tree.column(index,width = heading[1])
            tree.heading(index,text = heading[0])
        tree.grid(row=0,column=0)

        values = df[["Record","Track","Lat","Lon","Time"]].values.tolist()
        print("values",values[:25])
        for index,val in enumerate(values):
            tree.insert("", "end", iid=str(index), values=val)
        tree.bind("<<TreeviewSelect>>", self.view_gps_point)

        ###
        ### timing points frame
        ###

        tree = ttk.Treeview(self.timingPointsFrame, columns=[0, 1, 2],show="headings")
        for index, heading in enumerate([("Index",45), ("Lat",110), ("Lon",110)]):
            tree.column(index, width=heading[1])
            tree.heading(index, text=heading[0])
        tree.grid(row=0, column=0,columnspan=4)
        tree.bind("<Button-1>", self.view_timing_point)

        self.direction = tkinter.IntVar()
        self.direction.set(0)

        tkinter.Radiobutton(self.timingPointsFrame,text = "Primary",variable=self.direction,value = 0, command=self.direction_changed,bg="white").grid(row=1,column=0,columnspan=3)
        tkinter.Radiobutton(self.timingPointsFrame, text="Secondary", variable=self.direction,value =1, command=self.direction_changed,bg="white").grid(row=2, column=0,columnspan=3)
        tkinter.Label(self.timingPointsFrame,text ="Enter TP No.",bg="white",anchor=tkinter.E).grid(row=3,column=0,sticky="nsew")
        tkinter.Entry(self.timingPointsFrame,state="disabled",width = 5).grid(row=3,column=1,sticky="nsew")
        tkinter.Button(self.timingPointsFrame,state="disabled",text="save",bg="white").grid(row=3,column=2,sticky="nsew")
        #tkinter.Button(self.timingPointsFrame, state="disabled", text="cancel").grid(row=3, column=3)
        tkinter.Button(self.timingPointsFrame,text="Add",command=self.activate_add_timing_point,width = 7).grid(row=4,column=0,columnspan=3)
        tkinter.Button(self.timingPointsFrame, text="Delete", command=self.delete_timing_point,width = 7).grid(row=5, column=0,columnspan=3)
        tkinter.Button(self.timingPointsFrame, text="Save", command=self.save_timing_points,width = 7).grid(row=6, column=0,columnspan=3)
        tkinter.Button(self.timingPointsFrame, text="Full Track", command=lambda:self.mapViewer.set_run_indices(None),width = 7).grid(row=7, column=0,columnspan=3)

        ###
        ### Failed runs frame
        ###
        scroll2 = tkinter.Scrollbar(self.failedRunsFrame)
        tree = ttk.Treeview(self.failedRunsFrame,columns=[0,1,2], show="headings", height=20,yscrollcommand=scroll2.set)
        for index,heading in enumerate([("Start",40),("End",40),("Timing Point",80)]):
            tree.column(index,width = heading[1])
            tree.heading(index,text = heading[0])
        tree.grid(row=0,column=0)
        tree.bind("<Button-1>",self.failed_runs_clicked)
        scroll2.config(command=tree.yview)
        scroll2.grid(row=0, column=1, sticky="ns")
        tkinter.Radiobutton(self.failedRunsFrame,text = "Primary",variable=self.direction,value = 0, command=self.direction_changed,bg="white").grid(row=1,column=0)
        tkinter.Radiobutton(self.failedRunsFrame, text="Secondary", variable=self.direction,value =1, command=self.direction_changed,bg="white").grid(row=2, column=0)
        for item in self.baseData[self.direction.get()][3]:
            tree.insert("","end",values = item)
        print(win.winfo_reqwidth(),win.winfo_reqheight())
        win.geometry('420x1000+0+0')
        win.update()

        screen_width = self.winfo_screenwidth()
        left = screen_width - 1500
        self.mapViewer = mapViewer.MapViewer(1500,1000,left)
        self.mapViewer.set_centre_point(self.centrePoint)
        routeName = self.routeListBox.get(tkinter.ACTIVE)
        route = self.routes[routeName]

        self.mapViewer.set_coords(coords)
        self.mapViewer.set_route(route)
        self.mapViewer.set_callback_function("notify change of point",self.receive_notification_of_point_click)
        self.mapViewer.set_callback_function("notify added timing point", self.receive_notification_of_timing_point_added)
        self.mapViewer.set_callback_function("notify window closed", self.receive_notification_track_window_closed)
        self.direction_changed()
        self.update_map_viewer()

    def full_track_view_closed(self,win):
        self.routeListBox.config(state=tkinter.NORMAL)
        if not self.mapViewer is None:
            self.mapViewer.on_close()
        win.destroy()

    def scroll_with_scale(self,event):
        print("scroll value is",event)
        tree = self.trackFrame.winfo_children()[0]
        tree.see(int(float(event)))
        #print("value of slider is",event.widget.get())

    def update_map_viewer(self):
        if not self.mapViewer is None:
            self.mapViewer.update()
            self.after(10,self.update_map_viewer)

    def direction_changed(self):
        print("direction is", self.direction.get())
        routeName = self.routeListBox.get(tkinter.ACTIVE)
        tps = self.routes[routeName].get_timing_points()
        tree = self.timingPointsFrame.winfo_children()[0]
        tree.delete(*tree.get_children())
        tree.configure(height=len(tps[self.direction.get()]))
        for tp in tps[self.direction.get()]:
            tree.insert("", "end", iid=tp[0], values=[tp[0], tp[2], tp[3]])
        ###
        ### set up failed runs tree
        ###
        tree = self.failedRunsFrame.winfo_children()[1]
        tree.delete(*tree.get_children())
        for item in self.baseData[self.direction.get()][3]:
            tree.insert("", "end", values=item)
        self.mapViewer.set_timing_points_to_display(self.direction.get())

    def view_timing_point(self,event,index=None):
        if not event  is None:
            curItem = event.widget.identify_row(event.y)
            print(curItem)
            print(event.widget.selection())
            if curItem != "":
                self.mapViewer.view_timing_point(int(curItem)-1)
        if not index is None:
            self.mapViewer.view_timing_point(index-1)

    def activate_add_timing_point(self):
        if self.addTimingPointFlag == False:
            self.addTimingPointFlag = True
            print("colour of button is",self.timingPointsFrame.winfo_children()[6].cget("bg"))
            self.timingPointsFrame.winfo_children()[6].config(bg="green")
            self.mapViewer.set_cursor("CURSOR_CROSSHAIR")
        else:
            self.addTimingPointFlag = False
            self.timingPointsFrame.winfo_children()[6].config(bg="SystemButtonFace")
            self.mapViewer.set_cursor("CURSOR_DEFAULT")

    def view_gps_point(self, event, index=None):
        print("index is", index, "event is", event)
        start = time.time() * 1000
        if not event is None:
            selection = event.widget.selection()[0]
            event.widget.focus(selection)
            if selection != "":
                if selection == 'I001':
                    selection = "1"
                index = int(selection)
        if not index is None:
            self.mapViewer.view_gps_point(index, redraw=True)
        tree = self.trackFrame.winfo_children()[0]
        scale = self.trackFrame.winfo_children()[1]
        tree.update_idletasks()
        scale.set(index)
        return "break"

    def save_timing_points(self):
        file = filedialog.asksaveasfilename()
        if file == "":
            return
        if not ".txt" in file:
            file = file + ".txt"
        routeName = self.routeListBox.get(tkinter.ACTIVE)
        route = self.routes[routeName]
        route.save_timing_points(file)

    def delete_timing_point(self):
        tree = self.timingPointsFrame.winfo_children()[0]
        if len(tree.selection()) == 0:
            return
        selection = tree.selection()[0]
        tree.delete(selection)
        routeName = self.routeListBox.get(tkinter.ACTIVE)
        route = self.routes[routeName]
        route.delete_timing_point(self.direction.get(),selection)
        self.mapViewer.set_route(route)
        self.direction_changed()

    def receive_notification_of_point_click(self,index):
        print("received notification of change of point to",index)
        tree = self.trackFrame.winfo_children()[0]
        tree.selection_set(index)
        tree.see(tree.selection())
        self.trackTabs.select(1)
        tree.focus_set()

    def receive_notification_of_timing_point_added(self,data):
        ###
        ### receives a list of [newCoords,direction] from mapViewer
        ### passes them on when the Save button is pressed
        ###
        print("received notification of added timing point")
        self.timingPointsFrame.winfo_children()[4].config(state="normal",bg="red")
        self.timingPointsFrame.winfo_children()[5].config(state="normal",bg="red",command=lambda c = data:self.save_new_timing_point(c))
        self.timingPointsFrame.winfo_children()[6].config(state="disabled",bg="SystemButtonFace")

    def save_new_timing_point(self,data):
        newTPNo = self.timingPointsFrame.winfo_children()[4].get()
        try:
            newTPNo = int(newTPNo)
        except Exception as e:
            messagebox.showinfo(message="You must enter an integer for the timing point number")
            return
        routeName = self.routeListBox.get(tkinter.ACTIVE)
        route = self.routes[routeName]
        coords,direction = data
        route.add_timing_point(coords[0],coords[1],direction,newTPNo,reorder=True)
        self.timingPointsFrame.winfo_children()[6].config(bg="SystemButtonFace",state="normal")
        self.timingPointsFrame.winfo_children()[5].config(state="disabled",bg="white")
        self.timingPointsFrame.winfo_children()[4].config(state="disabled",bg="white")
        self.mapViewer.set_cursor("CURSOR_DEFAULT")
        self.addTimingPointFlag = False
        self.mapViewer.set_route(route)
        self.direction_changed()

    def failed_runs_clicked(self,event):
        widget = event.widget
        print(event.x,event.y)
        row = widget.identify_row(event.y)
        col = widget.identify_column(event.x)
        print("row,col",row,col)
        print(widget.identify_element(event.x,event.y))
        print(widget.set(row,col))
        try:
            runIndices = (int(widget.set(row,"#1")),int(widget.set(row,"#2")))
        except Exception as e:
            print(e,type(e))
            runIndices = None
        print("run indices are",runIndices)
        self.mapViewer.set_run_indices(runIndices)
        #if col  == "#3":
            #self.view_timing_point(None,index=int(widget.set(row,col)))
        #else:
            #self.view_gps_point(None,index=int(widget.set(row,col)))

    def receive_notification_track_window_closed(self):
        self.mapViewer = None
        self.routeListBox.config(state=tkinter.NORMAL)

#win = mainWindow()
#win.mainloop()
