
import tkinter
import tkinter.ttk as ttk
import tkinter.font as font
import datetime
import mapmanager
import openpyxl
import win32com.client
from PIL import Image,ImageDraw,ImageTk,ImageFont
from tkinter import filedialog
from tkinter import messagebox
import threading
import queue
import copy
import time
import os



class mainWindow(tkinter.Tk):

    def __init__(self):
        super(mainWindow, self).__init__()
        self.routes = {}
        self.tracsisBlue = "#%02x%02x%02x" % (20, 27, 77)
        ttk.Style().configure(".",foreground = self.tracsisBlue,background = "white",weight="bold")
        #self.bind("<<finished>>",self.received)
        self.selectedRoute = None
        self.progress = None
        self.trackData = None
        self.unitsVar = tkinter.IntVar()
        self.q = queue.Queue()
        self.discardedRuns = []
        self.baseData = [] ### store the returned data, ready to display either normal or discarded runs
        self.normalRuns=[]
        self.colours = ["black","red","gold","sea green","deep sky blue"]
        self.processSingleTrackFunction = None
        self.loadRoutesFunction = None
        self.getDateFunction = None
        self.getSpeedFunction = None
        self.previousLegIndex = -1 # this keeps track of the leg that is displayed yellow, so we dont have to redraw the whole track each time
        self.primaryTrackList = []
        self.secondaryTrackList = []
        self.trackWindow = None
        self.mapMan = None
        self.mapImage = None
        self.wm_title("JoPro - Journey Time Software")
        self.state("zoomed")
        self.primaryTrees = []
        self.secondaryTrees = []
        self.configure(bg="white")

        ###
        ### set up menu bar
        ###
        self.menubar = tkinter.Menu(self)
        menu = tkinter.Menu(self.menubar,tearoff = 0)
        menu.add_command(label = "Load TPs",command = self.displayRoutes)
        menu.add_separator()
        menu.add_command(label="Export",command=self.spawn_excel_window)
        self.menubar.add_cascade(label="File",menu = menu)
        self.config(menu=self.menubar)
        menu = tkinter.Menu(self.menubar,tearoff = 0)
        menu.add_command(label ="Journey Time Settings",command = self.spawn_settings_window)
        #menu.add_separator()
        #menu.add_command(label = "Excel Settings",command = self.spawn_excel_window)
        self.menubar.add_cascade(label = "Settings",menu = menu)
        self.thumbnail = None

        labelFont  = font.Font(family='Helvetica', size=16, weight='bold')

        ttk.Style().configure("Treeview.Heading", background="black")

        topFrame = tkinter.Frame(self,width = 1020,height = 1000, bg="white")
        subFrame = tkinter.Frame(topFrame, width=100, height=900, bg="white")
        self.routeListBox = tkinter.Listbox(master = subFrame,height = 10,width=17,relief=tkinter.SUNKEN,borderwidth =5,bg="white")
        #self.TPListBox = tkinter.Listbox(master=topFrame, height=15, relief=tkinter.SUNKEN, borderwidth=5,bg="white")
        buttonFrame = tkinter.Frame(subFrame,height = 100,bg="white")

        tkinter.Button(buttonFrame,text="+",font = labelFont,bg="white",width=2,command=lambda :self.change_zoom(1)).grid(row = 0,column = 0,sticky="ne",pady = (250,5))
        tkinter.Button(buttonFrame, text="-", font=labelFont,bg="white",width=2,command=lambda :self.change_zoom(-1)).grid(row=1, column=0,sticky="se",pady = 10)
        buttonFrame.grid(row = 1,column=0,pady = (200,0),sticky="e")

        self.discardedLabel= tkinter.Label(subFrame,text="Discarded\n Runs \n 0",font=labelFont,bg= "white",fg = self.tracsisBlue)
        self.discardedLabel.grid(row=1,column = 0,sticky= "n",pady = 10)
        self.discardedLabel.bind("<Double-Button-1>",self.display_discarded_runs)

        self.thumbnailCanvas = tkinter.Canvas(master = topFrame,width = 805,height = 805,relief=tkinter.SUNKEN,borderwidth =5,bg="white")
        self.routeListBox.bind('<Double-Button-1>', self.runProcess)
        self.routeListBox.bind('<<ListboxSelect>>', self.showTPs)
        tkinter.Label(topFrame,text = "Route",font=labelFont,justify=tkinter.LEFT,bg="white",fg = self.tracsisBlue).grid(row=0,column=0,padx=(10,0), pady= (0,11),sticky="nw")
        self.mapLabel = tkinter.Label(topFrame, text="Map", font=labelFont,justify=tkinter.LEFT,bg="white",fg = self.tracsisBlue)
        self.mapLabel.grid(row=0, column=1,padx=(10,0),pady= 0,sticky="nw")

        subFrame.grid(row=1,column=0, sticky="nw",padx=(40,30))
        self.routeListBox.grid(row=0, column=0, pady=2, padx=0, sticky="nw")


        self.thumbnailCanvas.grid(row=1, column=1, pady=0, padx=10, sticky="nw")

        topFrame.grid(row=0, column=0,sticky="nw", pady=30,padx=0)
        topFrame.grid_propagate(False)


        frame = tkinter.Frame(self,bg="white") ### outer frame, that contains everything in this segment

        ###
        ### set up the logo
        ###
        self.logoCanvas = tkinter.Canvas(frame, width=200, height=150, borderwidth=0, highlightthickness=0,relief='ridge',bg="white")
        img = Image.open("tracsis Logo.jpg")
        img = img.resize((int(img.width / 2), int(img.height / 2)), Image.ANTIALIAS)
        self.logo = ImageTk.PhotoImage(img)
        self.logoCanvas.create_image(5, 5, image=self.logo, anchor=tkinter.NW)
        self.logoCanvas.grid(row=3, column=0, pady=5, padx=10, sticky="ne")

        ###
        ### set up the notebook, and frames for each page in the notebook, add them to the notebook
        ###

        self.tabs = ttk.Notebook(frame)
        self.matrixFrame = tkinter.Frame(self, relief=tkinter.SUNKEN, borderwidth=5, bg="white")
        self.matrixFrame2 = tkinter.Frame(self, relief=tkinter.SUNKEN, borderwidth=5, width=800, height=794, bg="white")
        self.tabs.add(self.matrixFrame, text="Primary")
        self.tabs.add(self.matrixFrame2, text="Secondary")

        ###
        ### set up a canvas and a scrollbar for each page of the notebook
        ###

        canvas = tkinter.Canvas(self.matrixFrame,width = 800,height = 764,bg="white")
        hbar = tkinter.Scrollbar(self.matrixFrame, orient=tkinter.HORIZONTAL,command=canvas.xview)
        canvas.config(xscrollcommand=hbar.set,scrollregion=(0,0,810,0))
        canvas.grid(row=0,column=0)
        hbar.grid(row=1,column=0,sticky="we")
        canvas.xview_moveto(0)
        canvas.yview_moveto(0)

        canvas2 = tkinter.Canvas(self.matrixFrame2, width=800, height=764,bg="white")
        hbar2 = tkinter.Scrollbar(self.matrixFrame2, orient=tkinter.HORIZONTAL, command=canvas.xview)
        canvas2.config(xscrollcommand=hbar2.set, scrollregion=(0, 0, 810, 0))
        canvas2.grid(row=0, column=0)
        hbar2.grid(row=1, column=0, sticky="we")
        canvas2.xview_moveto(0)
        canvas2.yview_moveto(0)

        ###
        ### add a frame to each canvas
        ###

        innerFrame = tkinter.Frame(canvas,bg="white")
        canvas.create_window((0,0),window=innerFrame,anchor="nw")
        innerFrame2 = tkinter.Frame(canvas2, bg="white")
        canvas2.create_window((0, 0), window=innerFrame2, anchor="nw")


        self.bind_all("<<NotebookTabChanged>>", self.tabChanged)



        self.journeyLabel = tkinter.Label(frame, text="Journey Time Summary", font=labelFont,justify=tkinter.LEFT,bg="white",fg = self.tracsisBlue)
        self.journeyLabel.grid(row=0, column=0,pady= (29,11),sticky="nw")

        self.tabs.grid(row=1,column=0,sticky="nw")
        ttk.Style().configure("Treeview", background="light grey")
        frame.grid(row=0,column = 1, pady=0, padx=(40,0),sticky="ne")
        self.loadSettings()
        self.update()
        print("canvas size is",canvas.winfo_width(),canvas.winfo_height(),canvas.winfo_reqwidth(),canvas.winfo_reqheight())
        print("frame size is", innerFrame.winfo_width(), innerFrame.winfo_height(), innerFrame.winfo_reqwidth(),innerFrame.winfo_reqheight())


    def scroll_data_window(self,event):
        print("event",event,event.widget.get())
        left, right = (event.widget.get())
        thumbsize = right - left
        f = event.widget.fraction(event.x, event.y)
        if f < left:
            f = f - (thumbsize / 2)
        print("fraction is",f)
        self.dataWindow.xview_moveto(f)
        self.dataWindow.xview_moveto(f)

    def tabChanged(self,event):
        frame = None
        print("tab changed, widget type is", type(event.widget))
        ch = self.nametowidget(event.widget.select())
        print("child of tab is",type(ch))
        for child in ch.winfo_children():
            w = self.nametowidget(child)
            if type(w) == tkinter.Canvas:
                frame= w.winfo_children()[0]
        print(type(frame))
        if self.baseData == []:
            return

        text = self.discardedLabel.cget("text")
        print("calling widget",event.widget)
        print("index",self.tabs.index(self.tabs.select()))
        index = self.tabs.index(self.tabs.select())

        #print("setting scrollregion to",self.primaryTrees[0].winfo_width())
        if "Discarded" in text:
            noOfRuns = len(self.baseData[index][2])
            self.discardedLabel.configure(text="Discarded\n Runs\n" + str(noOfRuns))
        else:
            noOfRuns = len(self.baseData[index][0])
            self.discardedLabel.configure(text="Normal\n Runs\n" + str(noOfRuns))

    def display_discarded_runs(self,event):
        text = self.discardedLabel.cget("text")
        index = self.tabs.index(self.tabs.select())

        if "Discarded" in text:
            noOfRuns = len(self.baseData[index][0])
            self.discardedLabel.configure(text="Normal\n Runs\n" + str(noOfRuns))
            prim = copy.deepcopy(self.baseData[0])
            sec = copy.deepcopy(self.baseData[1])
            dataToDisplay = [[prim[2], prim[1]], [sec[2], sec[1]]]
            self.display_data(dataToDisplay)
        else:
            noOfRuns = len(self.baseData[index][2])
            self.discardedLabel.configure(text="Discarded\n Runs\n" + str(noOfRuns))
            prim = list(self.baseData[0])
            sec = list(self.baseData[1])
            dataToDisplay = [[prim[0], prim[1]], [sec[0], sec[1]]]
            self.display_data(dataToDisplay)

    def display_data(self, result):



        if result is None:
            # print("Stopping progress now -----")
            self.stopProgress()
            return
        primary = result[0]
        secondary = result[1]
        # print("Secondary is ",secondary)
        self.displayPrimary(primary)
        self.displaySecondary(secondary)
        # ("Stopping progress now")
        self.stopProgress()

    def startProgress(self,msg):

        self.progressWin = tkinter.Toplevel(self,width = 200,height = 200)
        x = int(self.winfo_screenwidth()/2 - 100)
        y = int(self.winfo_screenheight() / 2 - 100)
        self.progressWin.attributes("-topmost",True)
        tkinter.Label(self.progressWin,text = msg).grid(row=0,column = 0,padx = 20,pady= 20)
        self.progress = ttk.Progressbar(self.progressWin, orient="horizontal", length=200, mode="indeterminate")
        print("width height",str(self.progressWin.winfo_width()) + "x" + str(self.progressWin.winfo_height()))
        print("progressbar is",type(self.progress))
        self.progress.grid(row=1,column = 0,padx = 20,pady= 20)
        self.progress.start(10)
        print(str(self.progressWin.winfo_width()) + "x" + str(self.progressWin.winfo_height()) + "+" + str(x) + "+" + str(y))
        self.progressWin.geometry("+" + str(x) + "+" + str(y))
        self.progressWin.lift()

    def stopProgress(self):
        #print("in self.stop progress")
        print("progressbar is", type(self.progress))
        if self.progress is None:
            pass
        else:
            self.progress.stop()
            self.progress =None
        self.progressWin.destroy()

    def showTPs(self,event):
        ### Get the current selection in the routes List box, and display the TPS for that route in the TPS list box
        ###
        #self.TPListBox.delete(0,tkinter.END)
        routeName = self.routeListBox.get(self.routeListBox.curselection())
        if routeName in self.routes:
            self.mapLabel.configure(text="Map - " + routeName)
            self.thumbnail = ImageTk.PhotoImage(self.routes[routeName].get_map())
            #self.routes[routeName].get_map().show()
            #print("img size is",self.thumbnail.width(),self.thumbnail.height())
            self.thumbnailCanvas.create_image(10, 10, image=self.thumbnail, anchor=tkinter.NW)
            self.mapMan = self.routes[routeName].getMapManager()
            #for tps in self.routes[routeName].get_timing_points():
                #for point in tps:
                    #self.TPListBox.insert(tkinter.END,point)

    def key_pressed(self,event):
        if event.widget in self.primaryTrees[:3]:
            item = self.primaryTrees[0].selection()[0]
            index = self.primaryTrees[0].index(item)
            for tree in self.primaryTrees[:3]:
                tree.delete(item)
                self.map_closed()
            del self.primaryTrackList[index]
        if event.widget in self.secondaryTrees[:3]:
            item = self.secondaryTrees[0].selection()[0]
            index = self.secondaryTrees[0].index(item)
            for tree in self.secondaryTrees[:3]:
                tree.delete(item)
                self.map_closed()
            del self.secondaryTrackList[index]
        self.check_tags(self.primaryTrees, self.primaryTrackList)
        self.check_tags(self.secondaryTrees, self.secondaryTrackList)

    def map_closed(self):
        if self.trackWindow is None:
            pass
        else:
            self.trackWindow.destroy()
            self.trackWindow = None

    def settings_closed(self):
        self.saveSettings()
        self.settingsWindow.destroy()
        self.settingsWindow = None
        self.check_tags(self.primaryTrees,self.primaryTrackList)
        self.check_tags(self.secondaryTrees,self.secondaryTrackList)

    def excel_settings_closed(self):
        self.excelWindow.destroy()

    def export(self):
        file = filedialog.asksaveasfilename()
        print(file)
        if file == "":
            messagebox.showinfo(message="no file name entered, exiting Export")
            return
        for child in self.excelWindow.winfo_children():
            widget = self.nametowidget(child)
            if type(widget) == tkinter.Checkbutton:
                if widget.cget("text") == "Primary":
                    if widget.get() == 1:
                        print("yes")

        dir = self.entryValues[4].get()
        am = self.entryValues[1].get()
        ip = self.entryValues[2].get()
        pm = self.entryValues[3].get()
        primaryDirection = self.cbox.get()
        if primaryDirection == "North":
            primaryDirection = "Northbound"
            secondaryDirection = "Southbound"
        if primaryDirection == "South":
            primaryDirection = "Southbound"
            secondaryDirection = "Northbound"
        if primaryDirection == "East":
            primaryDirection = "Eastbound"
            secondaryDirection = "Westbound"
        if primaryDirection == "West":
            primaryDirection = "Westbound"
            secondaryDirection = "Eastbound"
        if primaryDirection == "Clockwise":
            secondaryDirection = "Anticlockwise"
        if primaryDirection == "Anticlockwise":
            secondaryDirection = "Clockwise"
        if self.check1.get() == 1:
            TPs = [(x[2], x[3]) for x in self.selectedRoute.get_timing_points()[0]]
            folder = os.path.dirname(os.path.abspath(__file__))
            folder = os.path.join(folder, "Runs\\")
            if not os.path.exists(folder):
                os.makedirs(folder)
            for fileName in os.listdir(folder):
                file_path = os.path.join(folder, fileName)
                try:
                    if os.path.isfile(file_path):
                        os.unlink(file_path)
                except Exception as e:
                    print(e)
            self.save_track_as_image(self.primaryTrackList)
            self.export_to_excel(self.primaryTrees,TPs,self.primaryTrackList,file + " " + primaryDirection)


        if self.check2.get() == 1:
            folder = os.path.dirname(os.path.abspath(__file__))
            print(folder)
            folder = os.path.join(folder, "Runs\\")
            print(folder)
            if not os.path.exists(folder):
                os.makedirs(folder)
            for fileName in os.listdir(folder):
                file_path = os.path.join(folder, fileName)
                try:
                    if os.path.isfile(file_path):
                        os.unlink(file_path)
                except Exception as e:
                    print(e)
            TPs = [(x[2], x[3]) for x in self.selectedRoute.get_timing_points()[1]]
            self.save_track_as_image(self.secondaryTrackList, direction="s")
            self.export_to_excel(self.secondaryTrees, TPs, self.secondaryTrackList,file + " " + secondaryDirection)

    def export_to_excel(self,trees,timingPoints,trackList,filename):
        AMRuns = 0
        IPRuns = 0
        PMRuns = 0
        runsList = []
        am1 = self.entryValues[0].get()
        am2 = self.entryValues[1].get()
        ip1 = self.entryValues[2].get()
        ip2 = self.entryValues[3].get()
        pm1 = self.entryValues[4].get()
        pm2 = self.entryValues[5].get()
        if self.selectedRoute is None:
            return
        img = Image.open("tracsis Logo.jpg")
        imgSmall = img.resize((184,65),Image.ANTIALIAS)
        excelImageSmall = openpyxl.drawing.image.Image(imgSmall)
        excelImage = openpyxl.drawing.image.Image(img)
        self.startProgress("Exporting to Excel")
        self.progressWin.update()
        surveyDate = self.getDateFunction(1)
        wb = openpyxl.load_workbook("Template.xlsm",keep_vba=True)
        sheets = wb.get_sheet_names()
        print("first sheet is",sheets[0])
        #wb.get_sheet_by_name(sheets[0]).add_image(excelImage,"B3")

        ###
        ### put big logo on front sheet
        ##

        sheet = wb.get_sheet_by_name(sheets[0])
        img = Image.open("tracsis Logo.jpg")
        imgSmall = img.resize((368, 130), Image.ANTIALIAS)
        excelImageSmall = openpyxl.drawing.image.Image(imgSmall)
        sheet.add_image(excelImageSmall, "B3")

        ###
        ### put little logos on the other sheets
        ###

        for sht in sheets[1:-1]:
            sheet = wb.get_sheet_by_name(sht)
            img = Image.open("tracsis Logo.jpg")
            imgSmall = img.resize((184, 65), Image.ANTIALIAS)
            excelImageSmall = openpyxl.drawing.image.Image(imgSmall)
            sheet.add_image(excelImageSmall,"A1")
        try:
            sheet = wb.get_sheet_by_name('Maps')
            for i in range(0, 8,2):
                sheet.cell(row=7+int(i/2),column=16).value = int(self.entryValues[8 + i].get())
                sheet.cell(row=7 +int(i/2), column=18).value = int(self.entryValues[8 + i+1].get())
            sheet.cell(row=11, column=16).value = self.entryValues[16].get() + "+"
        except Exception as e:
            print(e)
        try:
            sheet = wb.get_sheet_by_name('Temp')
        except Exception as e:
            print(e)
            return
        tpCount = len(timingPoints)
        for i,child in enumerate(trees[0].get_children()):
            self.progress.step()
            self.progressWin.update()
            flag = False
            #print(self.journeyTimesTree.item(child)["values"])
            #print(trees[0].item(child)["values"][1],type(trees[0].item(child)["values"][1]))
            if datetime.datetime.strptime(trees[0].item(child)["values"][1],"%H:%M:%S").time() >= datetime.datetime.strptime(am1,"%H:%M").time() and \
               datetime.datetime.strptime(trees[0].item(child)["values"][1],"%H:%M:%S").time() <= datetime.datetime.strptime(am2,"%H:%M").time():
                #print(trees[0].item(child)["values"][1],"is an AM Run")
                AMRuns+=1
                flag = True
                #print("adding",child,"to am runs")
                runsList.append(child)
            if datetime.datetime.strptime(trees[0].item(child)["values"][1],"%H:%M:%S").time() >= datetime.datetime.strptime(ip1, "%H:%M").time() and \
               datetime.datetime.strptime(trees[0].item(child)["values"][1],"%H:%M:%S").time() <= datetime.datetime.strptime(ip2, "%H:%M").time():
                #print(trees[0].item(child)["values"][1], "is an IP Run")
                IPRuns += 1
                runsList.append(child)
                #print("adding", child, "to ip runs")
                flag = True
            if datetime.datetime.strptime(trees[0].item(child)["values"][1],"%H:%M:%S").time() >= datetime.datetime.strptime(pm1, "%H:%M").time() and \
               datetime.datetime.strptime(trees[0].item(child)["values"][1],"%H:%M:%S").time() <= datetime.datetime.strptime(pm2,"%H:%M").time():
                #print(trees[0].item(child)["values"][1], "is an PM Run")
                PMRuns += 1
                runsList.append(child)
                #print("adding", child, "to pm runs")
                flag = True
        rawData = []
        print("runslist is",runsList)
        for i,run in enumerate(runsList):
            self.progress.step()
            self.progressWin.update()
            for j in range(tpCount):
                if self.check4.get() == 1:
                    ### add on an hour
                    sheet.cell(row=2 + i, column=1 + j).value =  datetime.datetime.strftime(datetime.datetime.strptime(trees[0].item(run)["values"][j + 1],"%H:%M:%S")+datetime.timedelta(hours=1),"%H:%M:%S")
                else:
                    sheet.cell(row=2+i,column = 1 + j).value = trees[0].item(run)["values"][j + 1]
            print(trackList[i][0])
            print(self.getDateFunction(trackList[i][0]))
            sheet.cell(row = 2+i,column = 1 + tpCount).value = self.getDateFunction(trackList[i][0])
        #for i,child in enumerate(trees[1].get_children()):
            for j in range(tpCount):
                sheet.cell(row=2 + i, column=j + tpCount + 2).value = trees[1].item(run)["values"][j + 1]
        #for i,child in enumerate(trees[2].get_children()):
            for j in range(tpCount):
                sheet.cell(row=2 + i, column=j + (2 * tpCount) + 2).value = trees[2].item(run)["values"][j + 1]


        sheet["A1"] = AMRuns
        sheet["B1"] = self.entryValues[0].get()
        sheet["C1"] = self.entryValues[1].get()
        sheet["D1"] = IPRuns
        sheet["E1"] = self.entryValues[2].get()
        sheet["F1"] = self.entryValues[3].get()
        sheet["G1"] = PMRuns
        sheet["H1"] = self.entryValues[4].get()
        sheet["I1"] = self.entryValues[5].get()
        sheet["J1"] = tpCount
        sheet["K1"] = surveyDate
        folder = os.path.dirname(os.path.abspath(__file__))
        print(folder)
        folder = os.path.join(folder, "Runs\\")
        sheet["L1"]=folder



        if self.getTrack != None:
            self.mapMan = mapmanager.MapManager(640, 640, 12, timingPoints[0], timingPoints)
            self.trackData = self.getTrack(trackList[1])
        offset=trackList[0][0]
        for i,t in enumerate(trackList[0][:-1]):
            self.progress.step()
            self.progressWin.update()
            #print(track[0][i],track[0][i+1],"offset is",offset)
            #print("dist between tps is",self.trackData["legDist"][self.trackList[0][i]-offset:self.trackList[0][i+1]-offset].sum())
            values = trees[5].item(trees[5].get_children()[0])["values"][1:]
            print("values are",values)
            sheet.cell(row=1,column = 13+i).value = values[i]
        lats = self.trackData["Lat"].tolist()
        lons = self.trackData["Lon"].tolist()
        pathData = list(zip(lats, lons))
        #print("Track data is ",pathData)
        image = self.mapMan.get_map_with_path(timingPoints,pathData)
        excelImage = openpyxl.drawing.image.Image(image)
        sheet = wb.get_sheet_by_name('Location - Distance')
        sheet.add_image(excelImage,"B13")




        if self.check3.get() == 1:
            ###
            ### dump raw data to the excel sheet
            ###
            for i, data in enumerate(trackList):
                try:
                    data = self.getTrack(data)
                    data["Track No"] = "Track " + str(runsList[i])
                    data[["Track No", "Record", "Time", "Lat", "Lon", "legTime", "legSpeed"]].apply(
                        lambda x: rawData.append(x.tolist()), axis=1)
                    sheet = wb.get_sheet_by_name('Raw Data')
                    for i, row in enumerate(rawData):
                        self.progress.step()
                        self.progressWin.update()
                        for j, item in enumerate(row):
                            sheet.cell(row=i + 1, column=j + 1).value = item
                except Exception as e:
                    print("PHOOO")
                    pass
                    ### total hack, if we have deleted some runs and try to dump the data, when we look for track x
        else:
            try:
                sht = wb.get_sheet_by_name("Raw Data")
                wb.remove_sheet(sht)
            except Exception as e:
                print("PHOOO")
                pass  ### we tried to remove the raw data sheet, but it already didnt exist

        try:
            wb.save(filename +".xlsm")
            xl = win32com.client.Dispatch("Excel.Application")
            xl.Application.Visible = True
            xlsFile = os.path.realpath(filename + ".xlsx")
            filename = filename+".xlsm"
            print("trying to open workbook",os.path.realpath(filename),xlsFile)
            time.sleep(0.5)
            wb = xl.Workbooks.Open(Filename=os.path.realpath(filename), ReadOnly=1)
            #xl.Workbooks.Open(Filename=os.path.realpath("C:/Users/NWatson/PycharmProjects/JourneyTimes/blah" + ".xlsm"), ReadOnly=1)
            xl.Application.Run("formatfile")

        except PermissionError as e:
            messagebox.showinfo(message="cannot save file- " + filename + " workbook is already open, please close and run export again")
        except Exception as e:
            print("couldnt save",e)

        self.excel_settings_closed()
        self.stopProgress()

    def set_end_point(self,event):
        curItem = event.widget.identify_row(event.y)
       # print("currently viewing track", self.displayedTrackIndex)
        #print("clicked on index",event.widget.index(curItem))
        #print("clicked on ", curItem)
        if self.displayedTrackDirection == "primary":
            currentTrack = self.primaryTrackList[self.displayedTrackIndex]
            TPs = [(x[2], x[3]) for x in self.selectedRoute.get_timing_points()[0]]
            trees = self.primaryTrees
            trackList = self.primaryTrackList
        if self.displayedTrackDirection == "secondary":
            currentTrack = self.secondaryTrackList[self.displayedTrackIndex]
            TPs = [(x[2], x[3]) for x in self.selectedRoute.get_timing_points()[1]]
            trees = self.secondaryTrees
            trackList = self.secondaryTrackList

        index = event.widget.index(curItem) + currentTrack[0]
        if index <currentTrack[-2]:
            return 'break'
        #print("current track data is ", currentTrack, "index is", index)
        currentTrack[-1] = index
        result = self.processSingleTrackFunction(currentTrack[0],currentTrack[-1],TPs)
        #print("received back",result)
        trackList[self.displayedTrackIndex] = result[0]
        trackName = trees[0].item(trees[0].selection()[0], "values")
        # print("row data is",trackName)
        values = result[1]
        values.insert(0, trackName[0])
        trees[0].item(trees[0].selection()[0], values=values)
        values = result[2]
        values.insert(0, trackName[0])
        trees[1].item(trees[1].selection()[0], values=values)
        values = result[3]
        values.insert(0, trackName[0])
        trees[2].item(trees[2].selection()[0], values=values)
        self.check_tags(trees, trackList)
        event.widget.item(curItem, values=result[1])
        self.spawn_track_window(None)
        return 'break'

    def set_start_point(self,event):
        curItem = event.widget.identify_row(event.y)
        #print("currently viewing track", self.displayedTrackIndex)
       # print("clicked on index", event.widget.index(curItem))
        #print("clicked on ", curItem)
        if self.displayedTrackDirection == "primary":
            currentTrack = self.primaryTrackList[self.displayedTrackIndex]
            TPs = [(x[2], x[3]) for x in self.selectedRoute.get_timing_points()[0]]
            trees = self.primaryTrees
            trackList = self.primaryTrackList
        if self.displayedTrackDirection == "secondary":
            currentTrack = self.secondaryTrackList[self.displayedTrackIndex]
            TPs = [(x[2], x[3]) for x in self.selectedRoute.get_timing_points()[1]]
            trees = self.secondaryTrees
            trackList = self.secondaryTrackList

        index = event.widget.index(curItem) + currentTrack[0]
        if index >= currentTrack[1]:
            return 'break'
        #print("current track data is ", currentTrack, "index is", index)
        currentTrack[0] = index
        result = self.processSingleTrackFunction(currentTrack[0], currentTrack[-1], TPs)
        # print("received back",result)
        trackList[self.displayedTrackIndex] = result[0]
        trackName = trees[0].item(trees[0].selection()[0], "values")
        # print("row data is",trackName)
        values = result[1]
        values.insert(0, trackName[0])
        trees[0].item(trees[0].selection()[0], values=values)
        values = result[2]
        values.insert(0, trackName[0])
        trees[1].item(trees[1].selection()[0], values=values)
        values = result[3]
        values.insert(0, trackName[0])
        trees[2].item(trees[2].selection()[0], values=values)
        self.check_tags(trees, trackList)
        event.widget.item(curItem, values=result[1])
        self.spawn_track_window(None)
        return 'break'

    def spawn_track_window(self,event):
        if event is None:
            if self.displayedTrackDirection == "primary":
                index = self.displayedTrackIndex
                track = self.primaryTrackList[self.displayedTrackIndex]
                TPs = [(x[2], x[3]) for x in self.selectedRoute.get_timing_points()[0]]
            if self.displayedTrackDirection == "secondary":
                index = self.displayedTrackIndex
                track = self.secondaryTrackList[self.displayedTrackIndex]
                TPs = [(x[2], x[3]) for x in self.selectedRoute.get_timing_points()[1]]
        else:
            if event.widget in self.primaryTrees:
                #print("selection is",self.primaryTrees[0].selection())
                if self.primaryTrees[0].selection() == "":
                    return
                index =(self.primaryTrees[0].index(self.primaryTrees[0].selection()[0]))
                track = self.primaryTrackList[index]
                self.displayedTrackDirection = "primary"
                self.displayedTrackIndex = index
                TPs = [(x[2], x[3]) for x in self.selectedRoute.get_timing_points()[0]]
            if event.widget in self.secondaryTrees:
                #print("selection is", self.secondaryTrees[0].selection())
                if self.secondaryTrees[0].selection() == "":
                    return
                index = (self.secondaryTrees[0].index(self.secondaryTrees[0].selection()[0]))
                track = self.secondaryTrackList[index]
                self.displayedTrackDirection = "secondary"
                self.displayedTrackIndex = index
                TPs = [(x[2], x[3]) for x in self.selectedRoute.get_timing_points()[1]]
        if self.trackWindow == None:

            #scrollbar = tkinter.Scrollbar(frame).pack(side=tkinter.RIGHT, fill=tkinter.Y)
            self.displayedTrackIndex = index
            self.trackWindow = tkinter.Toplevel(self)
            self.trackWindow.protocol("WM_DELETE_WINDOW", self.map_closed)
            frame = tkinter.Frame(master=self.trackWindow)
            frame.grid(row=0,column=0)
            scroll =tkinter.Scrollbar(frame)
            self.trackTree = ttk.Treeview(frame,columns=(1,2,3,4,5,6),show="headings",height = 7,yscrollcommand=scroll.set)
            scroll.config(command = self.trackTree.yview)
            self.trackTree.grid(row=0,column=0)
            scroll.grid(row=0,column=1,sticky="ns")
            self.trackTree.column(1,width = 65)
            self.trackTree.column(2, width=130)
            self.trackTree.column(3, width=65)
            self.trackTree.column(4, width=65)
            self.trackTree.column(5, width=65)
            self.trackTree.column(6, width=65)
            self.trackTree.heading(1,text="Record")
            self.trackTree.heading(2, text="Time")
            self.trackTree.heading(3, text="Lat")
            self.trackTree.heading(4, text="Lon")
            self.trackTree.heading(5, text="Leg Time")
            self.trackTree.heading(6, text="Leg Speed")

            self.trackTree.bind("<<TreeviewSelect>>",self.draw_track)
            self.trackTree.bind("<Control-1>",self.set_start_point)
            self.trackTree.bind("<Control-Shift-1>", self.set_end_point)


            frame = tkinter.Frame(master=self.trackWindow,width = 800,height  =800)
            frame.grid(row=1,column=0)
            self.mapCanvas = tkinter.Canvas(frame,width = 800,height  =800)
            self.mapCanvas.pack(expand=tkinter.YES, fill=tkinter.BOTH)
        self.trackTree.delete(*self.trackTree.get_children())

        #self.trackTree.tag_configure("speedissue", background="yellow")
        self.mapCanvas.delete("all")
        self.previousLegIndex = -1

        if self.getTrack != None:
            self.mapMan = mapmanager.MapManager(640,640,12,TPs[0],TPs)
            #routeName = self.routeListBox.get(self.routeListBox.curselection())
            #for k, v in self.routes.items():
                #print("looking for ", routeName, v.name)
                #if v.name == routeName:
                    #self.mapMan = v.getMapManager()
            self.mapImage =  ImageTk.PhotoImage(self.mapMan.get_map())
            self.mapCanvas.create_image(0, 0, image = self.mapImage, anchor = tkinter.NW)
            self.trackData = self.getTrack(track)
            lats = self.trackData["Lat"].tolist()
            lons = self.trackData["Lon"].tolist()
            pathData = list(zip(lats,lons))
            #self.mapMan.get_map_with_path(primaryTPs,pathData)
            self.trackData[["Record","Time","Lat","Lon","legTime","legSpeed"]].apply(lambda x: self.trackTree.insert("","end",values = x.tolist()),axis  = 1)
            self.trackTree.selection_set(self.trackTree.get_children()[0])
            self.trackTree.focus_set()
            self.trackTree.focus(self.trackTree.get_children()[0])
            self.trackTree.see(self.trackTree.get_children()[0])
            self.draw_track(None)
        self.trackWindow.geometry("800x1000")

    def draw_track(self,event):
        if self.trackData is None:
            return
        #print(self.trackData.head())
        #print(self.trackData.tail())
        if len(self.trackTree.selection()) >0:
            index = (self.trackTree.index(self.trackTree.selection()[0]))
        else:
            index = -1
        if self.previousLegIndex == -1:
            speed = int(self.trackData.iloc[self.previousLegIndex]["legSpeed"])
            #print("speed is",speed,"legindex is",self.previousLegIndex)
            if speed > 75:
                speed  = 75
            self.trackData[:-1].apply(lambda row: self.draw_leg((row["Lat"], row["Lon"]), (row["latNext"], row["lonNext"]), row["legSpeed"]), axis=1)
        if index >=0:
            speed = float(self.trackData.iloc[self.previousLegIndex]["legSpeed"])
            #print("speed is",speed)
            #colour = "#%02x%02x%02x" % (self.colour_gradient["r"][speed], self.colour_gradient["g"][speed], self.colour_gradient["b"][speed])

            self.draw_leg(
                (self.trackData.iloc[self.previousLegIndex]["Lat"], self.trackData.iloc[self.previousLegIndex]["Lon"]),
                (self.trackData.iloc[self.previousLegIndex]["latNext"],
                 self.trackData.iloc[self.previousLegIndex]["lonNext"]),speed)
            speed = int(self.trackData.iloc[index]["legSpeed"])
            if speed > 74:
                speed = 74
            #print("speed is", speed)
            #colour = "#%02x%02x%02x" % (
            #self.colour_gradient["r"][speed], self.colour_gradient["g"][speed], self.colour_gradient["b"][speed])
            self.draw_leg((self.trackData.iloc[index]["Lat"],self.trackData.iloc[index]["Lon"]),(self.trackData.iloc[index]["latNext"],self.trackData.iloc[index]["lonNext"]),-1)
            self.previousLegIndex = index

    def save_track_as_image(self,trackList,direction="p"):
        for i,track in enumerate(trackList):
            try:
                if direction=="p":
                    #image = Image.open(self.mapMan.get_map()).convert('RGB')
                    image = self.mapMan.get_map().copy()
                    print("image is ",image,type(image))
                else:
                    image = self.mapMan.get_sec_map().copy()
                    #image = Image.open(self.mapMan.get_sec_map()).convert('RGB')
                    print("image is ", image, type(image))
            except Exception as e:
                print("image error",e)
                return
            #image1 = ImageTk.PhotoImage(image)
            fnt = ImageFont.truetype("arial", size=18)
            drawimage = ImageDraw.Draw(image)
            trackData = self.getTrack(track)
            trackData[:-1].apply(lambda row: self.draw_leg_on_image((row["Lat"], row["Lon"]), (row["latNext"], row["lonNext"]), row["legSpeed"],drawimage),axis=1)
            t = datetime.datetime.strftime(trackData.iloc[0]["Time"],"%H:%M:%S")
            drawimage.rectangle([0, 0, 100, 50], fill="white")
            drawimage.text((10, 10), text=t, font=fnt, fill="black")
            folder = os.path.dirname(os.path.abspath(__file__))
            folder = os.path.join(folder, "Runs\\")
            image.save(folder + "/track " + str(i+1) + ".jpg")

    def draw_leg_on_image(self,p1,p2,speed,drawImage):
        colours = [(0,0,0), (255,0,0), (255,215,0),(46,139,87), (0,191,255)]
        if speed == -1:
            colour = "white"
        else:
            flag = False
            for i in range(0, 8, 2):
                #print(i, speed, self.entryValues[8 + i].get(), self.entryValues[8 + i + 1].get())
                if speed >= float(self.entryValues[8 + i].get()) and speed <= float(self.entryValues[8 + i + 1].get()):
                    colour = colours[int(i / 2)]
                    flag = True
                    #print("colour is", colour)
            if speed > float(self.entryValues[16].get()):
                colour = colours[-1]
                flag = True
            if not flag:
                colour = "white"
        x, y = self.mapMan.get_coords(p1)
        drawImage.ellipse([x - 5, y - 5, x + 5, y + 5], fill=colour)
        x1, y1 = self.mapMan.get_coords(p2)
        drawImage.ellipse([x1 - 5, y1 - 5, x1 + 5, y1 + 5], fill=colour)
        drawImage.line([x, y, x1, y1], fill=colour, width=6)

    def draw_leg(self,p1, p2, speed):
        #print("in leg speed, speed is",speed)
        if speed == -1:
            colour = "white"
        else:
            flag = False
            for i in range(0, 8, 2):
                #print(i, speed, self.entryValues[8 + i].get(), self.entryValues[8 + i + 1].get())
                if speed >= float(self.entryValues[8 + i].get()) and speed <= float(self.entryValues[8 + i + 1].get()):
                    colour = self.colours[int(i / 2)]
                    flag = True
                    #print("colour is", colour)
            if speed > float(self.entryValues[16].get()):
                colour = self.colours[-1]
                flag = True
            if not flag:
                colour = "white"
        #print("colour is",colour)
        x, y = self.mapMan.get_coords(p1)
        self.mapCanvas.create_oval([x - 5, y - 5, x + 5, y + 5], fill=colour,width=0)
        #d.ellipse([x - 5, y - 5, x + 5, y + 5], fill=colour)
        x1, y1 = self.mapMan.get_coords(p2)
        self.mapCanvas.create_oval([x1 - 5, y1 - 5, x1 + 5, y1 + 5], fill=colour,width=0)
        #d.ellipse([x1 - 5, y1 - 5, x1 + 5, y1 + 5], fill=colour)

        self.mapCanvas.create_line(x, y, x1, y1, fill=colour, width=6, smooth=True, splinesteps=400)
        #self.mapCanvas.create_oval([x - 1, y - 1, x + 1, y + 1], fill="white",width=0)
        #self.mapCanvas.create_oval([x1 - 1, y1 - 1, x1 + 1, y1 + 1], fill="white",width=0)
        #d.line([x, y, x1, y1], fill=colour, width=6)
        #d.ellipse([x - 3, y - 3, x + 3, y + 3], fill="white")
        #d.ellipse([x1 - 3, y1 - 3, x1 + 3, y1 + 3], fill="white")

    def list_selected(self,event):
        if event.widget in self.primaryTrees[:3]:
            curItem = event.widget.identify_row(event.y)
            for tree in self.primaryTrees[:3]:
                tree.selection_set(curItem)
                tree.focus_set()
                tree.see(curItem)
        if event.widget in self.secondaryTrees[:3]:
            curItem = event.widget.identify_row(event.y)
            for tree in self.secondaryTrees[:3]:
                tree.selection_set(curItem)
                tree.focus_set()
                tree.see(curItem)
        return 'break'

    def receive_processed_data(self,result):
        ###
        ### result is a list of lists
        ### result[0] holds the primary direction runs
        ### result[1] holds the secondary direction runs
        ### each specific direction holds [data,distances,discarded runs]
        ###

        ###
        ### set up the discarded runs label to show how many discarded runs there are
        ## then set up baseData, which holds the results of the journey times.
        ###

        self.tabs.select(self.tabs.tabs()[0]) ## select the Primary tab
        self.baseData = result
        print("result is",result)
        if result == []:
            return
        prim = copy.deepcopy(self.baseData[0]) ### there was a reason why I deep copied, cant remember....
        sec = copy.deepcopy(self.baseData[1])
        noOfRuns = len(self.baseData[0][2])
        self.discardedLabel.configure(text="Discarded\n Runs\n" + str(noOfRuns))

        if result is None:
            # print("Stopping progress now -----")
            self.stopProgress()
            return
        self.displayPrimary(prim)
        self.displaySecondary(sec)
        self.stopProgress()

    def runProcess(self,event):
        for tree in self.primaryTrees:
            tree.forget()
            tree.destroy()
        for tree in self.secondaryTrees:
            tree.forget()
            tree.destroy()
        self.primaryTrees=[]
        self.secondaryTrees=[]
        self.primaryTrackList = []
        self.secondaryTrackList = []
        routeName = self.routeListBox.get(self.routeListBox.curselection())
        if not routeName in self.routes:
            self.mapLabel.configure(text="Map")
            self.journeyLabel.configure(text="Journey Time Summary")
            return

        fileList = list(filedialog.askopenfilenames(initialdir=dir))
        if fileList == []:
            self.mapLabel.configure(text="Map")
            self.journeyLabel.configure(text="Journey Time Summary")
            return
        self.selectedRoute = self.routes[routeName]
        threading.Thread(target=wrapper_function,args = (self.fun,self.routes[routeName],fileList)).start()
        self.startProgress("Loading and Processing Data")
        self.mapLabel.configure(text="Map - " + routeName)
        self.journeyLabel.configure(text="Journey Time Summary - " + routeName)
        return

    def spawn_excel_window(self):
        self.excelWindow = tkinter.Toplevel(self)
        self.excelWindow.protocol("WM_DELETE_WINDOW", self.excel_settings_closed)
        frame = tkinter.Frame(self.excelWindow)
        #self.entryValues = []
        self.labels = []
        for i in range(6):
            self.labels.append(tkinter.Label(frame,anchor = tkinter.CENTER))
            #self.entryValues.append(tkinter.StringVar())
        self.labels[0].config(text="AM Times")
        self.labels[1].config(text="IP Times")
        self.labels[2].config(text="PM Times")
        for i in range(3,6):
            self.labels[i].config(text="to")

        for i in range(3):

            self.labels[i].grid(row=i, column=0, padx=2, pady=2)
            tkinter.Entry(frame, textvariable=self.entryValues[2*i], width=7).grid(row=i, column=1, padx=2, pady=2)
            self.labels[3+i].grid(row=i, column=2, padx=5, pady=5)
            tkinter.Entry(frame, textvariable=self.entryValues[(2*i)+1], width=7).grid(row=i, column=3, padx=2, pady=2)
        tkinter.Label(frame,text = "Primary Direction").grid(row = 3,column = 0)
        #self.entryValues.append(tkinter.StringVar())
        self.cbox = ttk.Combobox(frame,values=["North","South", "East","West", "Clockwise","Anticlockwise"],width = 13)
        self.cbox.grid(row = 3 ,column = 1)
        self.cbox.current(0)
        frame.grid(padx=10, pady=10)
        self.check1 = tkinter.IntVar()
        tkinter.Checkbutton(frame,text = "Primary",variable = self.check1).grid(row = 4,column = 0)
        self.check2 = tkinter.IntVar()
        tkinter.Checkbutton(frame, text="Secondary",variable = self.check2).grid(row=4, column=1)
        self.check3 = tkinter.IntVar()
        tkinter.Checkbutton(frame, text="Export Raw Data",variable = self.check3).grid(row=4, column=2)
        self.check4 = tkinter.IntVar()
        tkinter.Checkbutton(frame, text="Add 1 Hour to Times", variable=self.check4).grid(row=5, column=1)
        self.check1.set(1)
        self.check2.set(1)
        self.check3.set(0)
        self.check4.set(0)
        tkinter.Button(frame,text = "Export",command = self.export).grid(row  = 6,column  =0,padx=10, pady=10)
        tkinter.Button(frame, text="Exit", command = self.excel_settings_closed).grid(row=6, column=1,padx=10, pady=10)

    def spawn_settings_window(self):
        self.settingsWindow = tkinter.Toplevel(self)
        self.settingsWindow.protocol("WM_DELETE_WINDOW", self.settings_closed)
        frame = tkinter.Frame(self.settingsWindow)
        self.labels = []
        for i in range(7):
            self.labels.append(tkinter.Label(frame))
        self.labels[0].config(text = "Warning Speed")
        self.labels[1].config(text="Average Duration Warning %")
        self.labels[2].config(text="Congested/Stopped")
        self.labels[3].config(text="Slow")
        self.labels[4].config(text="Free Flowing - slow")
        self.labels[5].config(text="Free Flowing")
        self.labels[6].config(text="Free Flowing - fast")
        for i in range(2):
            self.labels[i].grid(row = i,column  = 0,padx = 2,pady =2)
            tkinter.Entry(frame,textvariable = self.entryValues[6 +i],width = 5).grid(row = i,column = 1)
        for i in range(2,7):
            self.labels[i].grid(row = i,column = 0)
        for i in range(5):
            #print(i)
            tkinter.Entry(frame, textvariable=self.entryValues[8 + (i * 2)], width=5).grid(row=2 + i, column=1)
            tkinter.Label(frame,text = "to").grid(row = 2 + i,column = 2)
            tkinter.Entry(frame, textvariable=self.entryValues[8 + (i *2) + 1], width=5).grid(row=2 + i, column=3)
        #print(frame.winfo_children()[-1].winfo_name())
        frame.winfo_children()[-1].grid_forget()
        frame.winfo_children()[-2].grid_forget()
        #self.unitsVar = tkinter.IntVar()
        #tkinter.Button(frame, text="Save", command=self.saveSettings).grid(row=2, column=0, padx=10, pady=10)
        tkinter.Button(frame, text="Exit", command=self.settings_closed).grid(row=11, column=2, padx=10, pady=10)
        tkinter.Radiobutton(frame, text="Miles", variable=self.unitsVar, value=1).grid(
            row=11, column=0, padx=(50, 0))
        tkinter.Radiobutton(frame, text="Kilometres", variable=self.unitsVar, value=2).grid(
            row=11, column=1, padx=(50, 0))
        frame.grid(padx = 10,pady =10)
        self.unitsVar.trace("w",self.units_changed)

    def get_units(self):
        return self.unitsVar.get()

    def units_changed(self,*args):
        ###
        ### user changed units in settings, either in miles or kilometres
        ###

        for dir in self.baseData:
            distList = []
            for track in dir[0]:
                result = self.getSpeedFunction(track[0])
                track[2] = result[0]
                distList.append(result[1])
            if len(distList) != 0:
                distList = [round(sum(i) / len(distList), 3) for i in zip(*distList)]
            else:
                distList = []
            dir[1] = distList
        prim = copy.deepcopy(self.baseData[0])  ### there was a reason why I deep copied, cant remember....
        sec = copy.deepcopy(self.baseData[1])
        self.displayPrimary(prim)
        self.displaySecondary(sec)

    def check_tags(self,trees,trackList):

        ###
        ### calculate and display the average durations
        ###
        if trees is None or trees == [] or trackList is None:
            return
        l = []
        print(trees[1])
        for row in trees[1].get_children():
            l.append(trees[1].item(row)[
                         "values"])  # we now have a list of lists, each sublist is the duration times for a specific leg of a journey
        l = list(zip(*l))[
            1:]  # zip the list of lists together , so that durations for the same leg are in the same list, knock off first list as that just contains the track numbers
        #print(l)
        denom = len(l[0])
        l = [[datetime.datetime.strptime(t, "%H:%M:%S") for t in sublist] for sublist in
             l]  # convert the strings to datetimes
        l = ["0" + str(
            sum([datetime.timedelta(seconds=init.hour * 3600 + init.minute * 60 + init.second) for init in sublist],
                datetime.timedelta(0)) / denom).split(".")[0] for sublist in
             l]  # convert the datetimes to timedeltas, sum the timedeltas, get the average, convert to string, format the string removing decimal places etc
        l.insert(0, "Average")
        #print(l)
        trees[3].delete(trees[3].get_children())
        trees[3].insert("", "end", values=l)

        ###
        ### calculate and display the average speeds
        ###

        l = []
        for row in trees[2].get_children():
            l.append(trees[2].item(row)["values"])
        l = list(zip(*l))[
            1:]  # zip the list of lists together , so that speeds for the same leg are in the same list, knock off first list as that just contains the track numbers
        l = [[float(t) for t in sublist] for sublist in l]  # convert the strings to floats
        l = [round(sum(item) / len(item), 2) for item in l]
        #print(l)

        l.insert(0, "Average")
        trees[4].delete(trees[4].get_children())
        trees[4].insert("", "end", values=l)


        ###
        ### check the durations to see if any are over average by X%
        ###
        if trees == []:
            return
        l = [x for x in trees[3].item(trees[3].get_children()[0])["values"]]
        #print("l is",l)
        for child in trees[1].get_children():
            durationtags = []
            for i, cell in enumerate(trees[1].item(child)["values"][1:]):
                averageTime = datetime.datetime.strptime(l[i + 1], "%H:%M:%S")
                t = datetime.datetime.strptime(cell, "%H:%M:%S")
                diff = (t - averageTime).total_seconds()
                denom = averageTime.minute * 60 + averageTime.second
                if denom !=0:
                    if diff * 100 / denom > float(self.entryValues[7].get()):
                        durationtags = ["timeissue"]
                trees[1].item(child, tags=durationtags)
        ###
        ### check if any average speed is greater than X
        ###
        speedtags = ["speedissue"]
        for child in trees[2].get_children():
            for i, cell in enumerate(trees[2].item(child)["values"][1:]):
                #print("checking",cell,float(self.entryValues[6].get()),float(cell) >= float(self.entryValues[6].get()))
                if float(cell) >= float(self.entryValues[6].get()):
                    #print("setting tag")
                    trees[2].item(child,tags=speedtags)
                    break
                else:
                    #print("clearing tag")
                    trees[2].item(child, tags=[])


        ###
        ### check the journey times to see if any are have same end time
        ###
        l = []

        for child in trees[0].get_children():
            trees[0].item(child,tags = [])
            #print("child is ", child)
            #print(trees[0].item(child)["values"])
            l.append((child,trees[0].item(child)["values"][-1]))
        #print("list of finish times is ", l)
        l2 = [x[1] for x in l]
        l = [child for child, x in l if l2.count(x) > 1]
        #print("l is ", l)
        [trees[0].item(i , tags=("timeissue",)) for i in l]

    def displaySecondary(self, result):
        for tree in self.secondaryTrees:
            tree.forget()
            tree.destroy()


        if result is None:
            return
        #print("result in secondary is",result)
        if result[0] == []:
            return

        ###
        ### identify the correct frame and canvas to put the widgets into
        ###
        frame = None
        tabList = self.tabs.tabs()
        ch = self.nametowidget(tabList[1])
        for child in ch.winfo_children():
            w = self.nametowidget(child)
            if type(w) == tkinter.Canvas:
                canvas = w
                frame = w.winfo_children()[0]

        distances = result[1]
        result = result[0] # because result contains both track and distance data
        if len(result) > 0:

            #####
            #####   Set up the new tables
            #####

            for tree in self.secondaryTrees:
                tree.forget()
                tree.destroy()
            self.secondaryTrees = []
            labels = []
            width = 65
            cols = tuple(range(len(result[0][0]) + 1))
            totalWidth = len(cols) * width
            self.secondaryTrees.append(
                ttk.Treeview(master=frame, columns=cols, show="headings", height=8))  ### journeytimes tree
            self.secondaryTrees.append(
                ttk.Treeview(master=frame, columns=cols, show="headings", height=8))  ### durations tree
            self.secondaryTrees.append(
                ttk.Treeview(master=frame, columns=cols, show="headings", height=8))  ### speed tree

            for tree in self.secondaryTrees:  ### set up the first column in all the trees
                tree.column(0, width=width, anchor='center')
                tree.heading(0, text='Track')
                tree.bind("<BackSpace>", self.key_pressed)
                tree.bind("<Delete>", self.key_pressed)
                tree.bind("<Double-Button-1>", self.spawn_track_window)
                tree.bind("<Button-1>", self.list_selected)

            for tree in self.secondaryTrees:
                for i in range(1, len(result[0][0]) + 1):  ### set up column headings for journey times
                    tree.column(i, width=width, anchor='center')
            self.secondaryTrees[1].heading(cols[-1], text="Total")
            self.secondaryTrees[2].heading(cols[-1], text="Average")

            self.secondaryTrees.append(ttk.Treeview(master=frame, columns=cols, height=1,
                                              show="headings"))  ### average durations tree
            self.secondaryTrees[3].column(0, width=width, anchor='center')
            self.secondaryTrees[3].column(cols[-1], width=width, anchor="center")
            self.secondaryTrees.append(
            ttk.Treeview(master=frame, columns=cols, height=1, show="headings"))  #### average speed tree
            self.secondaryTrees[4].column(0, width=width, anchor='center')
            self.secondaryTrees[4].column(cols[-1], width=width, anchor="center")
            self.secondaryTrees.append(
                ttk.Treeview(master=frame, columns=cols, height=1,show="headings"))  #### distance tree
            self.secondaryTrees[5].column(0, width=width, anchor='center')
            self.secondaryTrees[5].column(cols[-1], width=width, anchor="center")

            for i in range(1, len(result[0][0]) + 1):
                self.secondaryTrees[4].column(i, width=width, anchor='center')
                self.secondaryTrees[0].column(i, width=width, anchor='center')
                self.secondaryTrees[0].heading(i, text='TP' + str(i))

            for i in range(1, len(result[0][0]) + 1):  ### set up column headings for journey times
                self.secondaryTrees[4].column(i, width=width, anchor='center')
                self.secondaryTrees[0].column(i, width=width, anchor='center')
                self.secondaryTrees[0].heading(i, text='TP' + str(i))

            for tree in self.secondaryTrees[1:3]:  ### set up column headings and widths for duration and speed tables
                for i in range(1, len(result[0][0])):
                    tree.column(i, width=width, anchor='center')
                    tree.heading(i, text='TP' + str(i) + ' - TP' + str(i + 1))

            for tree in self.secondaryTrees[3:6]:  ### set up column widths for average tables
                for i in range(1, len(result[0][0])):
                    tree.column(i, width=width, anchor='center')
                    tree.heading(i, text='TP' + str(i) + ' - TP' + str(i + 1))

            self.secondaryTrees[3].heading(0, text="")
            self.secondaryTrees[4].heading(0, text="")
            self.secondaryTrees[5].heading(0, text="Distance")
            self.secondaryTrees[3].heading(len(result[0][0]), text="Total")
            self.secondaryTrees[4].heading(len(result[0][0]), text="Total")
            self.secondaryTrees[5].heading(len(result[0][0]), text="Total")

            ####
            #### set up the tags to denote different colours of rows
            ####

            self.secondaryTrees[0].tag_configure("timeissue", background="green")
            self.secondaryTrees[1].tag_configure("timeissue", background="red")
            self.secondaryTrees[2].tag_configure("speedissue", background="yellow")

            ###
            ### grid up the trees

            self.secondaryTrees[0].grid(row=0, column=0, pady=10, padx=10)
            self.secondaryTrees[1].grid(row=1, column=0, pady=(0, 0), padx=10)
            self.secondaryTrees[2].grid(row=3, column=0, pady=(10, 0), padx=10)
            self.secondaryTrees[3].grid(row=2, column=0, pady=(10, 0), padx=10)
            self.secondaryTrees[4].grid(row=4, column=0, pady=(10, 10), padx=10)
            self.secondaryTrees[5].grid(row=5, column=0, pady=0, padx=10)

            ttk.Style().configure("Treeview", background="grey")

            self.secondaryTrackList = [x[0] for x in
                              result]  ## self.tracklist is a list of lists, each entry is a list of indexes into the dataframe, indicating when that track hit a timing point

            ####
            ####      Go through the results, add them to the various tables
            ####
            for i, r in enumerate(result):
                print("displaying",r)

                ### insert the journey times

                times, speeds = list(r[1]), list(r[2])
                times.insert(0,  str(i + 1))
                self.secondaryTrees[0].insert("", "end", iid=i+1, values=times)

                ### insert the Durations

                times = times[1:]
                l = []
                for j in range(len(times) - 1):
                    l.append(datetime.datetime.strptime(times[j + 1], "%H:%M:%S") - datetime.datetime.strptime(times[j],
                                                                                                               "%H:%M:%S"))
                s = sum(l, datetime.timedelta())
                l.insert(0,  str(i + 1))
                l.insert(cols[-1], s)  ### insert
                self.secondaryTrees[1].insert("", "end", iid=i+1, values=l)

                ### insert the speeds

                speeds.insert(0, str(i + 1))
                self.secondaryTrees[2].insert("", "end", iid=i+1, values=speeds)
            if totalWidth < 500:
                totalWidth = 500
            #self.geometry(str(totalWidth + 60) + "x900")

            ###
            ### calculate and display the average durations
            ###

            l = []
            for row in self.secondaryTrees[1].get_children():
                l.append(self.secondaryTrees[1].item(row)[
                             "values"])  # we now have a list of lists, each sublist is the duration times for a specific leg of a journey
            l = list(zip(*l))[
                1:]  # zip the list of lists together , so that durations for the same leg are in the same list, knock off first list as that just contains the track numbers
            #print(l)
            denom = len(l[0])
            l = [[datetime.datetime.strptime(t, "%H:%M:%S") for t in sublist] for sublist in l]  # convert the strings to datetimes
            l = ["0" + str(
                sum([datetime.timedelta(seconds=init.hour * 3600 + init.minute * 60 + init.second) for init in sublist],
                    datetime.timedelta(0)) / denom).split(".")[0] for sublist in
                 l]  # convert the datetimes to timedeltas, sum the timedeltas, get the average, convert to string, format the string removing decimal places etc
            l.insert(0, "Average")
            #print(l)
            self.secondaryTrees[3].insert("", "end", values=l)


            ###
            ### calculate and display the average speeds
            ###

            l = []
            for row in self.secondaryTrees[2].get_children():
                l.append(self.secondaryTrees[2].item(row)["values"])
            l = list(zip(*l))[
                1:]  # zip the list of lists together , so that speeds for the same leg are in the same list, knock off first list as that just contains the track numbers
            l = [[float(t) for t in sublist] for sublist in l]  # convert the strings to floats
            #print(l)
            l = [round(sum(item) / len(item), 2) for item in l]
            #print(l)
            l.insert(0, "Average")
            print(l)
            self.secondaryTrees[4].insert("", "end", values=l)
            self.check_tags(self.secondaryTrees,self.secondaryTrackList)

        ###
        ### display the distance data
        ###

        distances.insert(0,"")
        self.secondaryTrees[5].insert("","end",values=distances)
        print("width of secondary tree is",self.secondaryTrees[0].winfo_width())

    def displayPrimary(self,result):
        for tree in self.primaryTrees:
            tree.forget()
            tree.destroy()
        if result is None:
            return
        if result[0] ==[]:
            return


        ###
        ### identify the correct frame and canvas to put the widgets into
        ###
        frame = None
        tabList = self.tabs.tabs()
        ch = self.nametowidget(tabList[0])
        print("child of tab is", type(ch))
        for child in ch.winfo_children():
            w = self.nametowidget(child)
            if type(w) == tkinter.Canvas:
                canvas = w
                frame = w.winfo_children()[0]
        print(type(frame))
        print("canvas size is", canvas.winfo_width(), canvas.winfo_height(), canvas.winfo_reqwidth(),
              canvas.winfo_reqheight())
        print("frame size is", frame.winfo_width(), frame.winfo_height(), frame.winfo_reqwidth(),
              frame.winfo_reqheight())

        distances = result[1]
        result = result[0]
        print("result is",result)
        if len(result) > 0:

            #####
            #####   Set up the new tables
            #####
            style = ttk.Style()
            style.configure("Treeview.Heading.label", font='helvetica 24')
            #print(style.layout("Treeheading"))
            #ttk.Style.configure(style="Treeview.Heading", foreground='white',bg="blue")
            self.primaryTrees = []
            labels = []
            width = 65
            cols = tuple(range(len(result[0][0]) + 1))
            totalWidth = len(cols) * width
            self.primaryTrees.append(
                ttk.Treeview(master=frame, columns=cols, show="headings", height=8))  ### journeytimes tree
            self.primaryTrees.append(
                ttk.Treeview(master=frame, columns=cols, show="headings", height=8))  ### durations tree
            self.primaryTrees.append(
                ttk.Treeview(master=frame, columns=cols, show="headings", height=8))  ### speed tree

            for tree in self.primaryTrees:  ### set up the first column in all the trees
                tree.column(0, width=width, anchor='center')
                tree.heading(0, text='Track')
                tree.bind("<BackSpace>", self.key_pressed)
                tree.bind("<Delete>", self.key_pressed)
                tree.bind("<Double-Button-1>", self.spawn_track_window)
                tree.bind("<Button-1>", self.list_selected)

            for tree in self.primaryTrees:  ### set up the end columns for the 2nd two tables
                tree.column(cols[-1], width=width, anchor="center")
            self.primaryTrees[1].heading(cols[-1], text="Total")
            self.primaryTrees[2].heading(cols[-1], text="Average")

            self.primaryTrees.append(ttk.Treeview(master=frame, columns=cols, height=1,
                                                  show="headings"))  ### average durations tree
            self.primaryTrees[3].column(0, width=width, anchor='center')
            self.primaryTrees[3].column(cols[-1], width=width, anchor="center")
            self.primaryTrees.append(
                ttk.Treeview(master=frame, columns=cols, height=1, show="headings"))  #### average speed tree
            self.primaryTrees[4].column(0, width=width, anchor='center')
            self.primaryTrees[4].column(cols[-1], width=width, anchor="center")
            self.primaryTrees.append(
                ttk.Treeview(master=frame, columns=cols, height=1, show="headings"))  #### distance tree
            self.primaryTrees[5].column(0, width=width, anchor='center')
            self.primaryTrees[5].column(cols[-1], width=width, anchor="center")

            for i in range(1, len(result[0][0]) + 1):  ### set up column headings for journey times
                self.primaryTrees[4].column(i, width=width, anchor='center')
                self.primaryTrees[0].column(i, width=width, anchor='center')
                self.primaryTrees[0].heading(i, text='TP' + str(i))

            for tree in self.primaryTrees[1:3]:  ### set up column headings and widths for duration and speed tables
                for i in range(1, len(result[0][0])):
                    tree.column(i, width=width, anchor='center')
                    tree.heading(i, text='TP' + str(i) + ' - TP' + str(i + 1))

            for tree in self.primaryTrees[3:6]:  ### set up column headings and widths for duration and speed tables
                for i in range(1, len(result[0][0])):
                    tree.heading(i, text='TP' + str(i) + ' - TP' + str(i + 1))
                    tree.column(i, width=width, anchor='center')
            self.primaryTrees[3].heading(0,text="")
            self.primaryTrees[4].heading(0, text="")
            self.primaryTrees[5].heading(0, text="Distance")
            self.primaryTrees[3].heading(len(result[0][0]), text="Total")
            self.primaryTrees[4].heading(len(result[0][0]), text="Total")
            self.primaryTrees[5].heading(len(result[0][0]), text="Total")

            ####
            #### set up the tags to denote different colours of rows
            ####

            self.primaryTrees[0].tag_configure("timeissue",background = "green")
            self.primaryTrees[1].tag_configure("timeissue", background="red")
            self.primaryTrees[2].tag_configure("speedissue", background="yellow")

            ###
            ### grid up the trees
            self.primaryTrees[0].grid(row=0, column=0, pady=10, padx=10)
            self.primaryTrees[1].grid(row=1, column=0, pady=(0,0), padx=10)
            self.primaryTrees[2].grid(row=3, column=0, pady=(10,0), padx=10)
            self.primaryTrees[3].grid(row=2, column=0, pady=(10,0), padx=10)
            self.primaryTrees[4].grid(row=4, column=0, pady=(10,10), padx=10)
            self.primaryTrees[5].grid(row=5, column=0, pady=0, padx=10)

            ttk.Style().configure("Treeview", background="light grey")

            self.primaryTrackList = [x[0] for x in
                              result]  ## self.tracklist is a list of lists, each entry is a list of indexes into the dataframe, indicating when that track hit a timing point

            ####
            ####      Go through the results, add them to the various tables
            ####

            for i, r in enumerate(result):


                times, speeds = list(r[1]), list(r[2])
                times.insert(0,str(i + 1))
                self.primaryTrees[0].insert("", "end", iid=i+1, values=times)

                ### insert the Durations

                times = times[1:]
                l = []
                for j in range(len(times) - 1):
                    l.append(datetime.datetime.strptime(times[j + 1], "%H:%M:%S") - datetime.datetime.strptime(times[j],
                                                                                                               "%H:%M:%S"))
                s = sum(l, datetime.timedelta())
                l.insert(0, str(i + 1))
                l.insert(cols[-1], s)  ### insert
                self.primaryTrees[1].insert("", "end", iid=i+1, values=l)

                ### insert the speeds

                speeds.insert(0,str(i + 1))
                self.primaryTrees[2].insert("", "end", iid=i+1, values=speeds)

            if totalWidth < 500:
                totalWidth = 500
            #self.geometry(str(totalWidth + 60) + "x900")

            ###
            ### calculate and display the average durations
            ###

            l = []
            for row in self.primaryTrees[1].get_children():
                l.append(self.primaryTrees[1].item(row)[
                             "values"])  # we now have a list of lists, each sublist is the duration times for a specific leg of a journey
            l = list(zip(*l))[
                1:]  # zip the list of lists together , so that durations for the same leg are in the same list, knock off first list as that just contains the track numbers
            #print(l)
            denom = len(l[0])
            l = [[datetime.datetime.strptime(t, "%H:%M:%S") for t in sublist] for sublist in
                 l]  # convert the strings to datetimes
            l = ["0" + str(
                sum([datetime.timedelta(seconds=init.hour * 3600 + init.minute * 60 + init.second) for init in sublist],
                    datetime.timedelta(0)) / denom).split(".")[0] for sublist in
                 l]  # convert the datetimes to timedeltas, sum the timedeltas, get the average, convert to string, format the string removing decimal places etc
            l.insert(0, "Average")
            #print(l)
            self.primaryTrees[3].insert("", "end", values=l)


            ###
            ### calculate and display the average speeds
            ###

            l = []
            for row in self.primaryTrees[2].get_children():
                l.append(self.primaryTrees[2].item(row)["values"])
            l = list(zip(*l))[
                1:]  # zip the list of lists together , so that speeds for the same leg are in the same list, knock off first list as that just contains the track numbers
            l = [[float(t) for t in sublist] for sublist in l]  # convert the strings to floats
            #print(l)
            l = [round(sum(item) / len(item), 2) for item in l]
            #print(l)
            l.insert(0, "Average")
            self.primaryTrees[4].insert("", "end", values=l)
            self.check_tags(self.primaryTrees,self.primaryTrackList)

        ###
        ### display the distance data
        ###

        distances.insert(0, "")
        self.primaryTrees[5].insert("", "end", values=distances)
        canvas.configure(scrollregion=(0,0,frame.winfo_reqwidth(),794))
        #frame.configure(width=self.primaryTrees[0].winfo_width())
        print("canvas size is", canvas.winfo_width(), canvas.winfo_height(), canvas.winfo_reqwidth(),
              canvas.winfo_reqheight())

        print("frame size is", frame.winfo_width(), frame.winfo_height(), frame.winfo_reqwidth(),
          frame.winfo_reqheight())

    def OnMouseWheel(self, event):
        self.journeyTimesTree.yview("scroll", -event.delta, "units")
        self.durationsTree.yview("scroll", -event.delta, "units")
        # this prevents default bindings from firing, which
        # would end up scrolling the widget twice
        return "break"

    def pressed(self):
        self.tree.forget()
        self.tree = ttk.Treeview(master=self.matrixFrame, columns=["1","2","3"], show="headings")
        self.tree.grid(row=0, column=0)
        self.tree.columns = ["1","2","3"]
        self.tree.heading("1",text = "Phoo")
        self.tree.insert("","end",values = ("wew","terte","fan"))
        self.tree.insert("","end",values=("wew", "terte", "fan"))

    def setCallbackFunction(self,text,fun):
        if text == "Routes":
            self.loadRoutesFunction = fun
        if text =="process":
            self.fun = fun
        if text == "getTrack":
            self.getTrack = fun
        if text == "getDate":
            self.getDateFunction = fun
        if text == "singleTrack":
            self.processSingleTrackFunction = fun
        if text == "getSpeed":
            self.getSpeedFunction = fun

    def change_zoom(self,val):
        if len(self.routeListBox.curselection()) ==0:
            return
        routeName = self.routeListBox.get(self.routeListBox.curselection())

        for k, v in self.routes.items():
            print("looking for ", routeName,v.name)
            if v.name == routeName:

                v.update_zoom(val)
                self.thumbnail = ImageTk.PhotoImage(v.get_map())
                self.thumbnailCanvas.create_image(10, 10, image=self.thumbnail, anchor=tkinter.NW)

    def displayRoutes(self):

        #threading.Thread(target = self.fun,args=(1,2)).start()
        #self.startProgress("Loading Timing Points")
       # print("hurrah")
        #return

        #self.event_generate("<<finished>>")
        for tree in self.primaryTrees:
            tree.forget()
            tree.destroy()
        for tree in self.secondaryTrees:
            tree.forget()
            tree.destroy()
        self.routeListBox.delete(0,tkinter.END)
        self.routes = self.loadRoutesFunction()
        if self.routes is None:
            return
        #self.startProgress("Loading Timing Points")
        for k,v in self.routes.items():
            #print(v.name)
            tps = [(x[2],x[3]) for x in v.get_timing_points()[0]]
            mp = mapmanager.MapManager(640, 640, 11, tps[0], tps)
            v.add_map(mp.get_thumbnail())
            v.setMapManager(mp)
            self.routeListBox.insert(tkinter.END,v.name)
        #self.stopProgress()

    def loadSettings(self):
        self.entryValues = []
        for i in range(18):
            self.entryValues.append(tkinter.StringVar())
        with open('settings.txt', 'r') as f:
            count = 0
            for e in self.entryValues:
                text = f.readline()
                text = text.replace("\n", "")
                if text == "":
                    pass
                e.set(text)
                print("loading",text)
            self.unitsVar.set(int(f.readline().replace("\n","")))

    def saveSettings(self):
        with open('settings.txt', 'w') as f:
            for e in self.entryValues:
                print("writing ", e.get())
                if e.get() == "":
                    f.write("\n")
                else:
                    f.write(str(e.get()) + "\n")
            print("writing ", self.unitsVar.get())
            f.write(str(self.unitsVar.get()) + "\n")

def wrapper_function(fun,routeName,fileList):
    result = fun(routeName,fileList)
    #self.q.put(result)
    return

def test_fun():
    for  i in range(10000000):
        print(i)
    return


def hex_to_RGB(hex):
    ''' "#FFFFFF" -> [255,255,255] '''
    # Pass 16 to the integer function for change of base
    return [int(hex[i:i + 2], 16) for i in range(1, 6, 2)]

def RGB_to_hex(RGB):
    ''' [255,255,255] -> "#FFFFFF" '''
    # Components need to be integers for hex to make sense
    RGB = [int(x) for x in RGB]
    return "#" + "".join(["0{0:x}".format(v) if v < 16 else
                          "{0:x}".format(v) for v in RGB])

def color_dict(gradient):
  ''' Takes in a list of RGB sub-lists and returns dictionary of
    colors in RGB and hex form for use in a graphing function
    defined later on '''
  return {"hex":[RGB_to_hex(RGB) for RGB in gradient],
      "r":[RGB[0] for RGB in gradient],
      "g":[RGB[1] for RGB in gradient],
      "b":[RGB[2] for RGB in gradient]}

def linear_gradient(start_hex, finish_hex="#FFFFFF", n=10):
  ''' returns a gradient list of (n) colors between
    two hex colors. start_hex and finish_hex
    should be the full six-digit color string,
    inlcuding the number sign ("#FFFFFF") '''
  # Starting and ending colors in RGB form
  s = hex_to_RGB(start_hex)
  f = hex_to_RGB(finish_hex)
  #print(s,f)
  # Initilize a list of the output colors with the starting color
  RGB_list = [s]
  # Calcuate a color at each evenly spaced value of t from 1 to n
  for t in range(1, n):
    # Interpolate RGB vector for color at the current value of t
    curr_vector = [
      int(s[j] + (float(t)/(n-1))*(f[j]-s[j]))
      for j in range(3)
    ]
    # Add it to our list of output colors
    RGB_list.append(curr_vector)

  return color_dict(RGB_list)

#window = mainWindow()
#window.spawn_track_window(None)
#window.mainloop()
